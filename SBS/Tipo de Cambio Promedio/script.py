"""
Scraper - Tipo de Cambio Promedio SBS
Tabla: Cotizacion de Oferta y Demanda (rgTipoCambio)
Columnas: FECHA | CODIGO | MONEDA | Compra | Venta
"""

import os
import sys
import time
import random
import threading
import subprocess
from datetime import date, timedelta
from pathlib import Path

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.scrolledtext as scrolledtext

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, JavascriptException,
    StaleElementReferenceException
)

from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# CONSTANTES
# ============================================================

URL_BASE        = (
    "https://www.sbs.gob.pe/app/pp/sistip_portal/paginas/publicacion/"
    "tipocambiopromedio.aspx"
)
ID_DATE_WRAPPER  = "ctl00_cphContent_rdpDate"
ID_DATE_INPUT    = "ctl00_cphContent_rdpDate_dateInput"
ID_BTN_CONSULTAR = "ctl00_cphContent_btnConsultar"
ID_LBL_FECHA     = "ctl00_cphContent_lblFecha"
ID_GRID          = "ctl00_cphContent_rgTipoCambio"

TABLA_EXCEL    = "TipoCambioSBS"
_SEG_POR_DIA   = 13
_MAX_RETROCESO = 45

_ISO: dict[str, str] = {
    "Dólar de N.A.":          "USD",
    "Dólar Estadounidense":   "USD",
    "Dólar Australiano":      "AUD",
    "Dólar Canadiense":       "CAD",
    "Dólar de Hong Kong":     "HKD",
    "Dólar de Singapur":      "SGD",
    "Dólar Neozelandés":      "NZD",
    "Dólar de N.Z.":          "NZD",
    "Peso Chileno":            "CLP",
    "Peso Colombiano":         "COP",
    "Peso Mexicano":           "MXN",
    "Peso Argentino":          "ARS",
    "Real Brasileño":          "BRL",
    "Real Brasilero":          "BRL",
    "Libra Esterlina":         "GBP",
    "Yen Japonés":             "JPY",
    "Yuan Chino":              "CNY",
    "Renminbi":                "CNY",
    "Won Coreano":             "KRW",
    "Franco Suizo":            "CHF",
    "Euro":                    "EUR",
    "Corona Sueca":            "SEK",
    "Corona Noruega":          "NOK",
    "Corona Danesa":           "DKK",
    "Corona Checa":            "CZK",
    "Zloty Polaco":            "PLN",
    "Sol Peruano":             "PEN",
    "Nuevo Sol":               "PEN",
}

_NOMBRE_USD = "Dólar Estadounidense"
_FMT_DATE   = "DD/MM/YYYY"
_FMT_NUM    = "#,##0.0000"

# ============================================================
# FERIADOS PERUANOS
# ============================================================

def _pascua(year: int) -> date:
    """Domingo de Pascua — algoritmo de Butcher."""
    a = year % 19; b = year // 100; c = year % 100
    d = b // 4;  e = b % 4;  f = (b + 8) // 25;  g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4;  k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day   = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def _feriados_peru(year: int) -> frozenset[date]:
    """Feriados nacionales peruanos: fijos + Semana Santa."""
    pascua = _pascua(year)
    fijos = [
        (1,  1), (5,  1), (6,  7), (6, 29),
        (7, 23), (7, 28), (7, 29), (8,  6),
        (8, 30), (10, 8), (11, 1), (12, 8),
        (12, 9), (12, 25),
    ]
    return frozenset(
        [date(year, m, d) for m, d in fijos]
        + [pascua - timedelta(days=3),  # Jueves Santo
           pascua - timedelta(days=2)]  # Viernes Santo
    )

# ============================================================
# UTILES
# ============================================================

def _directorio_exe() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _nombre_archivo(fi: date, ff: date) -> str:
    return f"TipoCambio_{fi.strftime('%d%m%Y')}_{ff.strftime('%d%m%Y')}.xlsx"


def _dias_habiles(fi: date, ff: date) -> int:
    return sum(
        1 for i in range((ff - fi).days + 1)
        if (fi + timedelta(days=i)).weekday() < 5
    )


def _fmt_duracion(segundos: int) -> str:
    return f"~{max(5, round(segundos / 60 / 5) * 5)} min"


def _normalizar_moneda(nombre: str) -> str:
    return _NOMBRE_USD if nombre == "Dólar de N.A." else nombre


def _trimestre_anterior(hoy: date | None = None) -> tuple[date, date]:
    """Devuelve (inicio, fin) del trimestre calendario inmediatamente anterior."""
    if hoy is None:
        hoy = date.today()
    q = (hoy.month - 1) // 3   # 0=Q1 actual, 1=Q2, 2=Q3, 3=Q4
    if q == 0:
        return date(hoy.year - 1, 10, 1), date(hoy.year - 1, 12, 31)
    starts = {1: (1, 1),  2: (4, 1),  3: (7, 1)}
    ends   = {1: (3, 31), 2: (6, 30), 3: (9, 30)}
    return (date(hoy.year, *starts[q]),
            date(hoy.year, *ends[q]))


# ============================================================
# SCRAPING
# ============================================================

def _crear_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1280,800")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--log-level=3")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)

    drv = webdriver.Chrome(options=opts)
    drv.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
    )
    return drv


def _esta_bloqueado(driver: webdriver.Chrome) -> bool:
    """Detecta el iframe de bloqueo de Imperva/Incapsula (Error 15)."""
    return '<iframe id="main-iframe"' in driver.page_source


def _esperar_pagina_lista(driver: webdriver.Chrome, timeout: int = 25) -> bool:
    try:
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.ID, ID_DATE_INPUT))
        )
        return True
    except TimeoutException:
        return False


def _esperar_telerik(driver: webdriver.Chrome, timeout: int = 15) -> None:
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return typeof $find !== 'undefined';")
        )
    except (TimeoutException, JavascriptException):
        pass


def _set_fecha(driver: webdriver.Chrome, fecha: date) -> None:
    """Establece la fecha vía API Telerik; fallback a input directo si falla."""
    js = (
        f"try {{ var r=$find('{ID_DATE_WRAPPER}');"
        f" if(!r) return false;"
        f" r.get_dateInput().set_value(new Date({fecha.year},{fecha.month-1},{fecha.day}));"
        f" return true; }} catch(e){{ return false; }}"
    )
    try:
        ok = driver.execute_script(js)
    except JavascriptException:
        ok = False

    if not ok:
        from selenium.webdriver.common.keys import Keys
        try:
            campo = driver.find_element(By.ID, ID_DATE_INPUT)
            driver.execute_script("arguments[0].value='';", campo)
            campo.click()
            campo.send_keys(fecha.strftime("%d/%m/%Y"))
            campo.send_keys(Keys.TAB)
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                campo,
            )
            time.sleep(0.2)
        except NoSuchElementException:
            pass


def _esperar_actualizacion(driver: webdriver.Chrome, label_anterior: str, timeout: int = 12) -> None:
    """
    Espera cualquier cambio en el label de fecha.
    Feriados responden en ~1 s con fecha distinta, sin agotar el timeout.
    """
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.find_element(By.ID, ID_LBL_FECHA).text != label_anterior
        )
    except TimeoutException:
        pass


def _parsear_grid(driver: webdriver.Chrome) -> list[dict]:
    soup = BeautifulSoup(driver.page_source, "lxml")
    div  = soup.find("div", id=ID_GRID)
    if not div:
        return []
    tabla = div.find("table", class_="rgMasterTable")
    if not tabla or not tabla.find("thead"):
        return []

    headers = [
        th.get_text(strip=True)
        for th in tabla.find("thead").find_all("th")
        if th.get_text(strip=True)
    ]
    if not headers:
        return []

    filas = tabla.find_all("tr", class_=lambda c: c and any(x in c for x in ["rgRow", "rgAltRow"]))
    resultado = []
    for fila in filas:
        vals = [td.get_text(strip=True) for td in fila.find_all("td")]
        if any(v for v in vals if v and v != "\xa0") and len(headers) == len(vals):
            resultado.append(dict(zip(headers, vals)))
    return resultado


def _limpiar_num(v: str) -> float | None:
    if not v or v in ("\xa0", "-", "S/M", ""):
        return None
    try:
        return float(v.replace(",", "").strip())
    except ValueError:
        return None


def _recuperar_de_bloqueo(
    driver: webdriver.Chrome,
    log_fn,
    pausa_fn=None,
) -> bool:
    """
    Intenta recuperar la sesión tras un bloqueo de Imperva.
    Backoff: 45 s → 90 s → 180 s.  Retorna True si se recuperó.
    Si pausa_fn está definida, ofrece al usuario intervenir manualmente
    antes del tercer intento (útil para resolver captchas a mano).
    """
    esperas = [45, 90, 180]
    for i, espera in enumerate(esperas):
        log_fn(f"  [BLOQUEO] Esperando {espera}s antes de reintentar ({i+1}/3)...")
        time.sleep(espera)
        driver.get(URL_BASE)
        time.sleep(3)
        if _esperar_pagina_lista(driver, timeout=20):
            if not _esta_bloqueado(driver):
                _esperar_telerik(driver)
                log_fn("  [OK] Sesión recuperada.")
                return True
        # Antes del último intento, ofrecer intervención manual si hay callback
        if i == 1 and pausa_fn is not None:
            pausa_fn()   # bloquea hasta que el usuario confirme
            driver.get(URL_BASE)
            time.sleep(3)
            if _esperar_pagina_lista(driver, timeout=20) and not _esta_bloqueado(driver):
                _esperar_telerik(driver)
                log_fn("  [OK] Sesión recuperada tras intervención manual.")
                return True
    log_fn("  [ERROR] No fue posible recuperar la sesión. Fechas omitidas hasta el próximo reintento.")
    return False


def _consultar_fecha(
    driver: webdriver.Chrome,
    fecha: date,
    primera_vez: bool,
    log_fn,
) -> list[dict]:
    if primera_vez:
        driver.get(URL_BASE)
        time.sleep(2)
        if not _esperar_pagina_lista(driver):
            log_fn("  [ERROR] La página no cargó.")
            return []
        _esperar_telerik(driver)
        log_fn("  Página lista.")

    _set_fecha(driver, fecha)
    time.sleep(random.uniform(0.4, 0.9))

    try:
        label_antes = driver.find_element(By.ID, ID_LBL_FECHA).text
    except NoSuchElementException:
        label_antes = ""

    # Retry anti-stale: el UpdatePanel puede reemplazar el nodo entre la
    # búsqueda del botón y el click si un refresh previo termina tarde
    clicked = False
    for _ in range(3):
        try:
            btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, ID_BTN_CONSULTAR))
            )
            driver.execute_script("arguments[0].click();", btn)
            clicked = True
            break
        except StaleElementReferenceException:
            time.sleep(0.3)
        except (TimeoutException, NoSuchElementException):
            break

    if not clicked:
        return []   # el caller detectará el estado y decidirá

    _esperar_actualizacion(driver, label_antes)
    time.sleep(random.uniform(0.2, 0.5))

    if _esta_bloqueado(driver):
        return []   # señal al caller para activar recovery

    return _parsear_grid(driver)


def _buscar_semillas(
    driver: webdriver.Chrome,
    df_raw: pd.DataFrame,
    fecha_inicio: date,
    log_fn,
) -> dict[str, dict]:
    """
    Para monedas con Compra=NaN en su primera fila, busca el último valor real
    previo a fecha_inicio. Una consulta por fecha hábil cubre todas las pendientes.
    """
    semillas: dict[str, dict] = {}
    pendientes: set[str] = set()

    for moneda in df_raw["MONEDA"].unique():
        sub      = df_raw[df_raw["MONEDA"] == moneda].sort_values("FECHA")
        primera  = sub.iloc[0]
        # Buscar semilla solo cuando Compra (el valor más crítico) está ausente
        if pd.isna(primera["Compra"]):
            pendientes.add(moneda)
            semillas[moneda] = {"Compra": None, "Venta": None}

    if not pendientes:
        return {}

    log_fn(f"  Buscando valor base anterior para: {', '.join(sorted(pendientes))}...")

    fecha = fecha_inicio - timedelta(days=1)
    for _ in range(_MAX_RETROCESO):
        if not pendientes:
            break
        if fecha.weekday() >= 5:
            fecha -= timedelta(days=1)
            continue

        filas = _consultar_fecha(driver, fecha, False, lambda _: None)

        for fila in filas:
            moneda = _normalizar_moneda(fila.get("MONEDA", ""))
            if moneda not in pendientes:
                continue
            compra = _limpiar_num(fila.get("COMPRA (S/)", ""))
            venta  = _limpiar_num(fila.get("VENTA (S/)",  ""))
            if semillas[moneda]["Compra"] is None and compra is not None:
                semillas[moneda]["Compra"] = compra
            if semillas[moneda]["Venta"] is None and venta is not None:
                semillas[moneda]["Venta"] = venta
            if semillas[moneda]["Compra"] is not None:
                pendientes.discard(moneda)

        if filas:
            cubiertas = sorted(m for m in semillas if m not in pendientes)
            if cubiertas:
                log_fn(f"  Valor base ({fecha.strftime('%d/%m/%Y')}): {', '.join(cubiertas)}")

        fecha -= timedelta(days=1)

    return semillas


def scrape_rango(
    fecha_inicio: date,
    fecha_fin: date,
    log_fn,
    cancelar_fn,
    progreso_fn=None,
    estado_fn=None,
    pausa_manual_fn=None,   # callback sin args: bloquea hasta que el usuario confirme reintentar
) -> tuple[pd.DataFrame, dict]:
    acum        = []
    driver      = _crear_driver()
    primera_vez = True
    t0          = time.time()
    semillas    = {}

    feriados: set[date] = set()
    for y in range(fecha_inicio.year, fecha_fin.year + 1):
        feriados.update(_feriados_peru(y))

    try:
        fechas    = [fecha_inicio + timedelta(days=i)
                     for i in range((fecha_fin - fecha_inicio).days + 1)]
        total     = len(fechas)           # días calendario totales
        consultas = sum(1 for f in fechas if f.weekday() < 5 and f not in feriados)
        contador  = 0
        consultadas = 0

        bloqueado    = False    # True si la sesión está actualmente bloqueada
        n_consultadas = 0       # para la pausa periódica anti-detección

        for fecha in fechas:
            if cancelar_fn():
                log_fn("Operación cancelada.")
                break

            contador += 1
            es_fin_semana = fecha.weekday() >= 5
            es_feriado    = fecha in feriados

            if es_fin_semana or es_feriado:
                if progreso_fn:
                    progreso_fn(int(contador / total * 100))
                continue

            consultadas    += 1
            n_consultadas  += 1
            fecha_str       = fecha.strftime("%d/%m/%Y")

            if estado_fn:
                estado_fn(f"[{contador}/{total}]  {fecha_str}")

            # Pausa periódica cada 80 consultas para reducir señales a Imperva
            if n_consultadas > 1 and n_consultadas % 80 == 0:
                pausa = random.randint(8, 15)
                log_fn(f"  [PAUSA] Descanso de {pausa}s para reducir detección...")
                time.sleep(pausa)

            filas = _consultar_fecha(driver, fecha, primera_vez, log_fn)
            primera_vez = False

            # Si no hubo filas, puede ser bloqueo o fecha sin datos real
            if not filas:
                if _esta_bloqueado(driver) or not bloqueado:
                    # Solo intentar recovery si aún no estábamos en estado bloqueado
                    # o si la página confirma bloqueo ahora
                    if _esta_bloqueado(driver):
                        log_fn(f"  [BLOQUEO] Detectado en {fecha_str}.")
                        ok = _recuperar_de_bloqueo(driver, log_fn, pausa_manual_fn)
                        bloqueado = not ok
                        if ok:
                            # Reintentar la misma fecha tras la recuperación
                            filas = _consultar_fecha(driver, fecha, False, log_fn)
                    else:
                        bloqueado = False

            if filas:
                bloqueado = False
                for fila in filas:
                    moneda = _normalizar_moneda(fila.get("MONEDA", ""))
                    acum.append({
                        "FECHA":  fecha,
                        "CODIGO": _ISO.get(moneda, "???"),
                        "MONEDA": moneda,
                        "Compra": _limpiar_num(fila.get("COMPRA (S/)", "")),
                        "Venta":  _limpiar_num(fila.get("VENTA (S/)", "")),
                    })
                log_fn(f"✓  {fecha_str}")
            else:
                log_fn(f"⚠  {fecha_str}  (sin datos)")

            if progreso_fn:
                progreso_fn(int(contador / total * 100))

        if acum and not cancelar_fn():
            df_temp  = pd.DataFrame(acum)
            semillas = _buscar_semillas(driver, df_temp, fecha_inicio, log_fn)

    finally:
        try:
            driver.quit()
        except Exception:
            pass

    minutos = max(1, round((time.time() - t0) / 60))
    log_fn(f"Tiempo total: ~{minutos} min.")

    if not acum:
        return pd.DataFrame(), {}
    return pd.DataFrame(acum).reset_index(drop=True), semillas


def aplicar_ffill(
    df: pd.DataFrame,
    fi: date,
    ff: date,
    semillas: dict | None = None,
) -> pd.DataFrame:
    """
    Reindexa a todos los días del rango y aplica ffill por moneda.
    Inyecta fila semilla previa a fi para cubrir el bloque inicial sin datos.
    """
    if df.empty:
        return df

    todos_dias    = pd.date_range(fi, ff, freq="D")
    fecha_semilla = pd.Timestamp(fi) - timedelta(days=1)
    orden_moneda  = {m: i for i, m in enumerate(df["MONEDA"].unique())}
    partes        = []

    for moneda in df["MONEDA"].unique():
        sub = df[df["MONEDA"] == moneda].copy()
        sub = sub.set_index("FECHA")
        sub.index = pd.DatetimeIndex(sub.index)

        # Inyectar semilla si fue encontrada y tiene al menos Compra
        if semillas and moneda in semillas:
            seed = semillas[moneda]
            if seed.get("Compra") is not None:
                sub.loc[fecha_semilla] = {
                    "CODIGO": _ISO.get(moneda, "???"),
                    "MONEDA": moneda,
                    "Compra": seed["Compra"],
                    "Venta":  seed.get("Venta"),
                }

        rango_ext = pd.DatetimeIndex([fecha_semilla] + list(todos_dias))
        sub = sub.reindex(rango_ext.sort_values())
        sub["CODIGO"] = sub["CODIGO"].ffill().bfill()
        sub["MONEDA"] = sub["MONEDA"].ffill().bfill()
        # Solo rellenar si existe al menos un valor real; ausencias estructurales
        # de SBS (ej. Peso Chileno sin Venta) se dejan como NaN intencionalmente
        if sub["Compra"].notna().any():
            sub["Compra"] = sub["Compra"].ffill().bfill()
        if sub["Venta"].notna().any():
            sub["Venta"] = sub["Venta"].ffill().bfill()

        sub = sub.loc[todos_dias]
        sub.index.name = "FECHA"
        partes.append(sub.reset_index())

    out = pd.concat(partes, ignore_index=True)
    out["FECHA"] = out["FECHA"].dt.date
    out["_ord"]  = out["MONEDA"].map(orden_moneda)
    out = out.sort_values(["FECHA", "_ord"]).drop(columns="_ord").reset_index(drop=True)
    return out


# ============================================================
# EXPORTACION A EXCEL
# ============================================================

_COLOR_HDR = "1F4E79"
_COLOR_FNT = "FFFFFF"
_COLOR_ALT = "DDEBF7"


def _borde():
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def exportar_excel(df: pd.DataFrame, ruta: str) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="TipoCambio")

    cols       = list(df.columns)
    idx_fecha  = cols.index("FECHA")  + 1
    idx_compra = cols.index("Compra") + 1
    idx_venta  = cols.index("Venta")  + 1

    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font      = Font(bold=True, color=_COLOR_FNT, size=10)
        cell.fill      = PatternFill("solid", fgColor=_COLOR_HDR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _borde()
    ws.row_dimensions[1].height = 18

    for r, row in df.iterrows():
        er     = r + 2
        es_alt = r % 2 == 1
        for c, val in enumerate(row, 1):
            if not isinstance(val, (date, str)) and pd.isna(val):
                val = None
            cell = ws.cell(row=er, column=c, value=val)
            cell.border    = _borde()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if es_alt:
                cell.fill = PatternFill("solid", fgColor=_COLOR_ALT)
            if c == idx_fecha:
                cell.number_format = _FMT_DATE
                cell.alignment     = Alignment(horizontal="center", vertical="center")
            elif c in (idx_compra, idx_venta):
                cell.number_format = _FMT_NUM
                cell.alignment     = Alignment(horizontal="right", vertical="center")

    anchos = {"FECHA": 14, "CODIGO": 9, "MONEDA": 24, "Compra": 13, "Venta": 13}
    for c, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = anchos.get(col, 12)

    ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"
    tbl = Table(displayName=TABLA_EXCEL, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    wb.save(ruta)


# ============================================================
# INTERFAZ GRAFICA
# ============================================================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tipo de Cambio SBS - Extractor")
        self.resizable(False, False)
        self._cancelar  = False
        self._hilo      = None
        self._t_inicio  = 0.0
        self._construir_ui()
        self.protocol("WM_DELETE_WINDOW", self._cerrar_ventana)

    # ---- Construcción ----

    def _construir_ui(self):
        PAD = {"padx": 10, "pady": 6}
        fi_def, ff_def = _trimestre_anterior()

        # --- Rango de fechas ---
        frm_f = ttk.LabelFrame(self, text="Rango de fechas")
        frm_f.grid(row=0, column=0, sticky="ew", **PAD)

        self._vi_d = tk.StringVar(value=f"{fi_def.day:02d}")
        self._vi_m = tk.StringVar(value=f"{fi_def.month:02d}")
        self._vi_y = tk.StringVar(value=str(fi_def.year))
        self._vf_d = tk.StringVar(value=f"{ff_def.day:02d}")
        self._vf_m = tk.StringVar(value=f"{ff_def.month:02d}")
        self._vf_y = tk.StringVar(value=str(ff_def.year))

        ttk.Label(frm_f, text="Fecha inicio:").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        ei_d, ei_m, ei_y = self._campos_fecha(frm_f, 0, 1, self._vi_d, self._vi_m, self._vi_y)

        ttk.Label(frm_f, text="Fecha fin:").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        ef_d, ef_m, ef_y = self._campos_fecha(frm_f, 1, 1, self._vf_d, self._vf_m, self._vf_y)

        self._auto_avanzar(self._vi_d, 2, ei_m)
        self._auto_avanzar(self._vi_m, 2, ei_y)
        self._auto_avanzar(self._vi_y, 4, ef_d)
        self._auto_avanzar(self._vf_d, 2, ef_m)
        self._auto_avanzar(self._vf_m, 2, ef_y)

        self._var_info = tk.StringVar(value="")
        ttk.Label(frm_f, textvariable=self._var_info, foreground="#666666").grid(
            row=2, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 4))

        for v in (self._vi_d, self._vi_m, self._vi_y,
                  self._vf_d, self._vf_m, self._vf_y):
            v.trace_add("write", lambda *_: self._actualizar_info())
        self._actualizar_info()

        # --- Carpeta de salida ---
        frm_s = ttk.LabelFrame(self, text="Carpeta de salida")
        frm_s.grid(row=1, column=0, sticky="ew", **PAD)
        frm_s.columnconfigure(0, weight=1)

        self._var_carpeta = tk.StringVar(value=str(_directorio_exe()))
        self._ent_carpeta = ttk.Entry(frm_s, textvariable=self._var_carpeta)
        self._ent_carpeta.grid(row=0, column=0, sticky="ew", padx=(8, 2), pady=4)
        ttk.Button(frm_s, text="...", width=3,
                   command=self._elegir_carpeta).grid(row=0, column=1, padx=(0, 8), pady=4)

        self._var_prev = tk.StringVar(value="")
        ttk.Label(frm_s, textvariable=self._var_prev,
                  foreground="#2255AA", font=("Segoe UI", 8)).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 4))

        for v in (self._vi_d, self._vi_m, self._vi_y,
                  self._vf_d, self._vf_m, self._vf_y):
            v.trace_add("write", lambda *_: self._actualizar_prev())
        self._actualizar_prev()

        # --- Opciones ---
        frm_o = ttk.LabelFrame(self, text="Opciones")
        frm_o.grid(row=2, column=0, sticky="ew", **PAD)

        self._var_abrir = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_o, text="Abrir archivo al terminar",
                        variable=self._var_abrir).grid(
            row=0, column=0, sticky="w", padx=8, pady=4)

        self._var_abrir_carpeta = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm_o, text="Abrir carpeta al terminar",
                        variable=self._var_abrir_carpeta).grid(
            row=0, column=1, sticky="w", padx=8, pady=4)

        self._var_ffill = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            frm_o,
            text="Llenar datos faltantes en fines de semana y feriados por el último valor disponible",
            variable=self._var_ffill,
        ).grid(row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))

        # --- Progreso ---
        frm_l = ttk.LabelFrame(self, text="Progreso")
        frm_l.grid(row=3, column=0, sticky="ew", **PAD)
        frm_l.columnconfigure(0, weight=1)

        # Barra de estado: [contador/fecha] a la izquierda, [timer] a la derecha
        frm_st = ttk.Frame(frm_l)
        frm_st.grid(row=0, column=0, columnspan=2, sticky="ew", padx=6, pady=(4, 0))
        frm_st.columnconfigure(0, weight=1)

        self._var_estado = tk.StringVar(value="")
        ttk.Label(frm_st, textvariable=self._var_estado,
                  foreground="#1F4E79", font=("Segoe UI", 8)).grid(
            row=0, column=0, sticky="w")

        self._var_timer = tk.StringVar(value="")
        ttk.Label(frm_st, textvariable=self._var_timer,
                  foreground="#888888", font=("Consolas", 8)).grid(
            row=0, column=1, sticky="e")

        self._log = scrolledtext.ScrolledText(
            frm_l, height=11, width=70, state="disabled",
            font=("Consolas", 9), wrap="word",
        )
        self._log.grid(row=1, column=0, columnspan=2, padx=6, pady=(2, 2))

        ttk.Button(frm_l, text="Limpiar", width=8,
                   command=self._limpiar_log).grid(
            row=2, column=1, sticky="e", padx=8, pady=(0, 4))

        # --- Barra de progreso ---
        self._barra = ttk.Progressbar(self, mode="determinate", length=460, maximum=100)
        self._barra.grid(row=4, column=0, padx=10, pady=4, sticky="ew")

        # --- Botones ---
        frm_b = ttk.Frame(self)
        frm_b.grid(row=5, column=0, pady=8)
        self._btn_ej = ttk.Button(frm_b, text="Ejecutar", width=14, command=self._iniciar)
        self._btn_ej.pack(side="left", padx=8)
        self._btn_ca = ttk.Button(frm_b, text="Cancelar", width=14,
                                  command=self._solicitar_cancelar, state="disabled")
        self._btn_ca.pack(side="left", padx=8)

        self.columnconfigure(0, weight=1)

    def _campos_fecha(self, parent, row, col, vd, vm, vy):
        vcmd2 = (self.register(lambda s: s == "" or (s.isdigit() and len(s) <= 2)), "%P")
        vcmd4 = (self.register(lambda s: s == "" or (s.isdigit() and len(s) <= 4)), "%P")

        frm = ttk.Frame(parent)
        frm.grid(row=row, column=col, sticky="w", padx=8, pady=4)

        ent_d = ttk.Entry(frm, textvariable=vd, width=3, validate="key", validatecommand=vcmd2)
        ent_d.pack(side="left")
        ttk.Label(frm, text=" / ").pack(side="left")
        ent_m = ttk.Entry(frm, textvariable=vm, width=3, validate="key", validatecommand=vcmd2)
        ent_m.pack(side="left")
        ttk.Label(frm, text=" / ").pack(side="left")
        ent_y = ttk.Entry(frm, textvariable=vy, width=5, validate="key", validatecommand=vcmd4)
        ent_y.pack(side="left")

        return ent_d, ent_m, ent_y

    def _auto_avanzar(self, var: tk.StringVar, max_len: int, siguiente: ttk.Entry):
        def _check(*_):
            v = var.get()
            if len(v) == max_len and v.isdigit():
                siguiente.focus_set()
                siguiente.select_range(0, tk.END)
        var.trace_add("write", _check)

    # ---- UI helpers ----

    def _pfecha(self, vd, vm, vy) -> date | None:
        try:
            return date(int(vy.get()), int(vm.get()), int(vd.get()))
        except (ValueError, OverflowError):
            return None

    def _actualizar_info(self):
        fi = self._pfecha(self._vi_d, self._vi_m, self._vi_y)
        ff = self._pfecha(self._vf_d, self._vf_m, self._vf_y)
        if fi and ff:
            if fi > ff:
                self._var_info.set("⚠  La fecha inicio es posterior a la fecha fin")
            else:
                t   = (ff - fi).days + 1
                est = _fmt_duracion(_dias_habiles(fi, ff) * _SEG_POR_DIA)
                self._var_info.set(f"{t} días en el rango  |  Duración aprox.: {est}")
        else:
            self._var_info.set("")

    def _actualizar_prev(self):
        fi = self._pfecha(self._vi_d, self._vi_m, self._vi_y)
        ff = self._pfecha(self._vf_d, self._vf_m, self._vf_y)
        if fi and ff and fi <= ff:
            self._var_prev.set(f"Archivo: {_nombre_archivo(fi, ff)}")
        else:
            self._var_prev.set("")

    def _elegir_carpeta(self):
        c = filedialog.askdirectory(
            initialdir=self._var_carpeta.get() or str(_directorio_exe()),
            title="Seleccionar carpeta de salida",
        )
        if c:
            self._var_carpeta.set(c)

    def _log_msg(self, texto: str):
        def _w():
            self._log.configure(state="normal")
            self._log.insert("end", texto + "\n")
            self._log.see("end")
            self._log.configure(state="disabled")
        self.after(0, _w)

    def _set_estado(self, texto: str):
        self.after(0, lambda: self._var_estado.set(texto))

    def _limpiar_log(self):
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")

    def _set_progreso(self, v: int):
        self.after(0, lambda: self._barra.configure(value=v))

    def _tick_timer(self):
        if self._hilo and self._hilo.is_alive():
            elapsed = int(time.time() - self._t_inicio)
            m, s = divmod(elapsed, 60)
            self._var_timer.set(f"{m:02d}:{s:02d}")
            self.after(1000, self._tick_timer)

    def _set_controles(self, ejecutando: bool):
        ej = "disabled" if ejecutando else "normal"
        ca = "normal"   if ejecutando else "disabled"
        self._btn_ej.configure(state=ej)
        self._btn_ca.configure(state=ca)
        for w in (self._ent_carpeta,):
            w.configure(state="disabled" if ejecutando else "normal")
        if not ejecutando:
            self._barra.configure(value=0)
            self._set_estado("")

    def _solicitar_cancelar(self):
        self._cancelar = True
        self._log_msg(">>> Cancelando...")

    def _pedir_reintento_manual(self):
        """
        Llamado por scrape_rango cuando el backoff automático no logró recuperar la sesión.
        Muestra un diálogo y bloquea hasta que el usuario resuelva el captcha en Chrome
        y confirme que puede continuar, o elija cancelar.
        """
        continuar = [False]
        ev = threading.Event()

        def _mostrar():
            resp = messagebox.askyesno(
                "Bloqueo de seguridad",
                "La página SBS está bloqueando el acceso.\n\n"
                "En la ventana de Chrome, recarga la página manualmente "
                "y resuelve el captcha si aparece.\n\n"
                "Haz clic en Sí cuando la página esté lista para continuar, "
                "o en No para cancelar.",
            )
            continuar[0] = resp
            if not resp:
                self._cancelar = True
            ev.set()

        self.after(0, _mostrar)
        ev.wait()   # bloquea el hilo de scraping hasta que el usuario responda

    def _cerrar_ventana(self):
        if self._hilo and self._hilo.is_alive():
            if not messagebox.askyesno(
                "Proceso en curso",
                "Hay una consulta en ejecución. Cerrar ahora puede dejar "
                "Chrome abierto.\n¿Deseas cerrar de todas formas?",
            ):
                return
        self.destroy()

    # ---- Ejecución ----

    def _iniciar(self):
        fi = self._pfecha(self._vi_d, self._vi_m, self._vi_y)
        ff = self._pfecha(self._vf_d, self._vf_m, self._vf_y)

        if fi is None:
            messagebox.showerror("Fecha inválida", "La fecha de inicio no es válida.")
            return
        if ff is None:
            messagebox.showerror("Fecha inválida", "La fecha fin no es válida.")
            return
        if fi > ff:
            messagebox.showerror(
                "Rango inválido",
                f"La fecha inicio ({fi.strftime('%d/%m/%Y')}) es posterior "
                f"a la fecha fin ({ff.strftime('%d/%m/%Y')}).",
            )
            return

        carpeta = self._var_carpeta.get().strip()
        if not carpeta:
            messagebox.showerror("Error", "Especifica la carpeta de salida.")
            return

        cp = Path(carpeta)
        if not cp.exists():
            if messagebox.askyesno("Carpeta no existe",
                                   f"La carpeta no existe:\n{carpeta}\n¿Deseas crearla?"):
                cp.mkdir(parents=True, exist_ok=True)
            else:
                return

        ruta = str(cp / _nombre_archivo(fi, ff))

        # Conflicto de archivo existente
        if Path(ruta).exists():
            resp = messagebox.askyesnocancel(
                "Archivo ya existe",
                f"'{Path(ruta).name}' ya existe en la carpeta seleccionada.\n\n"
                "Sí       → Sobreescribir\n"
                "No       → Elegir otro nombre\n"
                "Cancelar → Cancelar la operación",
            )
            if resp is None:
                return
            if resp is False:
                nueva = filedialog.asksaveasfilename(
                    initialdir=str(cp),
                    initialfile=Path(ruta).name,
                    defaultextension=".xlsx",
                    filetypes=[("Excel", "*.xlsx")],
                    title="Guardar como",
                )
                if not nueva:
                    return
                ruta = nueva

        est = _fmt_duracion(_dias_habiles(fi, ff) * _SEG_POR_DIA)

        self._cancelar = False
        self._set_controles(ejecutando=True)
        self._set_progreso(0)
        self._var_timer.set("00:00")

        self._log_msg("=" * 55)
        self._log_msg(f"Inicio: {fi.strftime('%d/%m/%Y')}  |  Fin: {ff.strftime('%d/%m/%Y')}")
        self._log_msg(f"Duración estimada: {est}  |  Archivo: {Path(ruta).name}")
        self._log_msg("=" * 55)
        self._log_msg("Chrome se abrirá. No lo cierres durante el proceso.")

        self._t_inicio = time.time()
        self.after(1000, self._tick_timer)

        self._hilo = threading.Thread(
            target=self._ejecutar_hilo,
            args=(fi, ff, ruta),
            daemon=True,
        )
        self._hilo.start()

    def _ejecutar_hilo(self, fi: date, ff: date, ruta: str):
        try:
            df, semillas = scrape_rango(
                fi, ff,
                log_fn=self._log_msg,
                cancelar_fn=lambda: self._cancelar,
                progreso_fn=self._set_progreso,
                estado_fn=self._set_estado,
                pausa_manual_fn=self._pedir_reintento_manual,
            )

            if self._cancelar:
                self._log_msg("Proceso cancelado. Archivo no generado.")
                return

            if df.empty:
                self._log_msg("[WARN] No se obtuvieron datos.")
                self.after(0, lambda: messagebox.showwarning(
                    "Sin datos",
                    "No se obtuvieron datos del sitio SBS.\n"
                    "Verifica que Chrome no esté siendo bloqueado.",
                ))
                return

            if self._var_ffill.get():
                self._log_msg("Expandiendo a todos los días del rango...")
                df = aplicar_ffill(df, fi, ff, semillas=semillas)

            self._log_msg("Generando archivo Excel...")
            exportar_excel(df, ruta)
            self._log_msg(f"Filas exportadas: {len(df)}")
            self._log_msg(f"Archivo: {ruta}")
            self._log_msg("Proceso completado.")
            self._set_progreso(100)
            self._set_estado("Completado.")

            if self._var_abrir.get():
                self.after(0, lambda: _abrir_archivo(ruta))
            if self._var_abrir_carpeta.get():
                self.after(0, lambda: _abrir_carpeta(str(Path(ruta).parent)))

        except Exception as exc:
            self._log_msg(f"[ERROR] {exc}")
            self.after(0, lambda: messagebox.showerror("Error", str(exc)))
        finally:
            self.after(0, lambda: self._set_controles(ejecutando=False))


# ============================================================
# UTILES DE SISTEMA
# ============================================================

def _abrir_archivo(ruta: str):
    try:
        if sys.platform == "win32":
            os.startfile(ruta)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", ruta])
        else:
            subprocess.Popen(["xdg-open", ruta])
    except Exception:
        pass


def _abrir_carpeta(ruta: str):
    try:
        if sys.platform == "win32":
            subprocess.Popen(["explorer", ruta])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", ruta])
        else:
            subprocess.Popen(["xdg-open", ruta])
    except Exception:
        pass


if __name__ == "__main__":
    app = App()
    app.mainloop()