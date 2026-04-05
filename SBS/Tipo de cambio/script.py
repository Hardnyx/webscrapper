"""
Scraper - Tipo de Cambio Promedio SBS
Tabla: Cotizacion de Oferta y Demanda (rgTipoCambio)
Columnas: FECHA | CODIGO | MONEDA | Compra | Venta
"""

import os
import sys
import time
import threading
import subprocess
from datetime import date, timedelta
from pathlib import Path

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.scrolledtext as scrolledtext

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, JavascriptException
)

from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# CONSTANTES
# ============================================================

URL_BASE         = (
    "https://www.sbs.gob.pe/app/pp/sistip_portal/paginas/publicacion/"
    "tipocambiopromedio.aspx"
)
ID_DATE_WRAPPER  = "ctl00_cphContent_rdpDate"
ID_DATE_INPUT    = "ctl00_cphContent_rdpDate_dateInput"
ID_BTN_CONSULTAR = "ctl00_cphContent_btnConsultar"
ID_LBL_FECHA     = "ctl00_cphContent_lblFecha"
ID_GRID          = "ctl00_cphContent_rgTipoCambio"

TABLA_EXCEL      = "TipoCambioSBS"
_SEG_POR_DIA     = 13
_MAX_RETROCESO   = 45

# Mapeo de nombres SBS a código ISO — incluye variantes ortográficas y monedas futuras
_ISO: dict[str, str] = {
    "Dólar de N.A.":          "USD",
    "Dólar Estadounidense":   "USD",
    "Dólar Australiano":      "AUD",
    "Dólar Canadiense":       "CAD",
    "Dólar de Hong Kong":     "HKD",
    "Dólar de Singapur":      "SGD",
    "Dólar Neozelandés":      "NZD",
    "Dólar de N.Z.":          "NZD",
    "Peso Chileno":            "CLP",
    "Peso Colombiano":         "COP",
    "Peso Mexicano":           "MXN",
    "Peso Argentino":          "ARS",
    "Real Brasileño":          "BRL",
    "Real Brasilero":          "BRL",
    "Libra Esterlina":         "GBP",
    "Yen Japonés":             "JPY",
    "Yuan Chino":              "CNY",
    "Renminbi":                "CNY",
    "Won Coreano":             "KRW",
    "Franco Suizo":            "CHF",
    "Euro":                    "EUR",
    "Corona Sueca":            "SEK",
    "Corona Noruega":          "NOK",
    "Corona Danesa":           "DKK",
    "Corona Checa":            "CZK",
    "Zloty Polaco":            "PLN",
    "Sol Peruano":             "PEN",
    "Nuevo Sol":               "PEN",
}

_NOMBRE_USD = "Dólar Estadounidense"
_FMT_DATE   = "DD/MM/YYYY"
_FMT_NUM    = "#,##0.0000"

# ============================================================
# FERIADOS PERUANOS
# ============================================================

def _pascua(year: int) -> date:
    """Domingo de Pascua — algoritmo de Butcher."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day   = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def _feriados_peru(year: int) -> frozenset[date]:
    """Feriados nacionales peruanos: fijos + Semana Santa."""
    pascua = _pascua(year)
    fijos = [
        (1,  1),  # Año Nuevo
        (5,  1),  # Día del Trabajo
        (6,  7),  # Batalla de Arica y Día de la Bandera
        (6, 29),  # San Pedro y San Pablo
        (7, 23),  # Día de la Fuerza Aérea del Perú
        (7, 28),  # Fiestas Patrias
        (7, 29),  # Fiestas Patrias
        (8,  6),  # Batalla de Junín
        (8, 30),  # Santa Rosa de Lima
        (10, 8),  # Combate de Angamos
        (11, 1),  # Día de Todos los Santos
        (12, 8),  # Inmaculada Concepción
        (12, 9),  # Batalla de Ayacucho
        (12, 25), # Navidad
    ]
    return frozenset(
        [date(year, m, d) for m, d in fijos]
        + [pascua - timedelta(days=3),   # Jueves Santo
           pascua - timedelta(days=2)]   # Viernes Santo
    )


# ============================================================
# UTILES
# ============================================================

def _directorio_exe() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _nombre_archivo(fi: date, ff: date) -> str:
    return f"TipoCambio_{fi.strftime('%d%m%Y')}_{ff.strftime('%d%m%Y')}.xlsx"


def _dias_habiles(fi: date, ff: date) -> int:
    return sum(
        1 for i in range((ff - fi).days + 1)
        if (fi + timedelta(days=i)).weekday() < 5
    )


def _fmt_duracion(segundos: int) -> str:
    return f"~{max(5, round(segundos / 60 / 5) * 5)} min"


def _normalizar_moneda(nombre: str) -> str:
    return _NOMBRE_USD if nombre == "Dólar de N.A." else nombre


# ============================================================
# SCRAPING
# ============================================================

def _crear_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1280,800")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--log-level=3")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)

    drv = webdriver.Chrome(options=opts)
    drv.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
    )
    return drv


def _esperar_pagina_lista(driver: webdriver.Chrome, timeout: int = 25) -> bool:
    try:
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.ID, ID_DATE_INPUT))
        )
        return True
    except TimeoutException:
        return False


def _esperar_telerik(driver: webdriver.Chrome, timeout: int = 15) -> None:
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return typeof $find !== 'undefined';")
        )
    except (TimeoutException, JavascriptException):
        pass


def _set_fecha(driver: webdriver.Chrome, fecha: date) -> None:
    """Establece la fecha vía API Telerik; fallback a input directo si falla."""
    js = (
        f"try {{ var r=$find('{ID_DATE_WRAPPER}');"
        f" if(!r) return false;"
        f" r.get_dateInput().set_value(new Date({fecha.year},{fecha.month-1},{fecha.day}));"
        f" return true; }} catch(e){{ return false; }}"
    )
    try:
        ok = driver.execute_script(js)
    except JavascriptException:
        ok = False

    if not ok:
        from selenium.webdriver.common.keys import Keys
        try:
            campo = driver.find_element(By.ID, ID_DATE_INPUT)
            driver.execute_script("arguments[0].value='';", campo)
            campo.click()
            campo.send_keys(fecha.strftime("%d/%m/%Y"))
            campo.send_keys(Keys.TAB)
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                campo,
            )
            time.sleep(0.2)
        except NoSuchElementException:
            pass


def _esperar_actualizacion(driver: webdriver.Chrome, label_anterior: str, timeout: int = 12) -> None:
    """
    Espera cualquier cambio en el label de fecha.
    Detecta feriados rápidamente (~1 s) porque el servidor responde con una
    fecha diferente en lugar de agotar el timeout completo.
    """
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.find_element(By.ID, ID_LBL_FECHA).text != label_anterior
        )
    except TimeoutException:
        pass


def _parsear_grid(driver: webdriver.Chrome) -> list[dict]:
    soup = BeautifulSoup(driver.page_source, "lxml")
    div  = soup.find("div", id=ID_GRID)
    if not div:
        return []
    tabla = div.find("table", class_="rgMasterTable")
    if not tabla:
        return []
    thead = tabla.find("thead")
    if not thead:
        return []

    headers = [th.get_text(strip=True) for th in thead.find_all("th") if th.get_text(strip=True)]
    if not headers:
        return []

    filas = tabla.find_all("tr", class_=lambda c: c and any(x in c for x in ["rgRow", "rgAltRow"]))
    resultado = []
    for fila in filas:
        vals = [td.get_text(strip=True) for td in fila.find_all("td")]
        if any(v for v in vals if v and v != "\xa0") and len(headers) == len(vals):
            resultado.append(dict(zip(headers, vals)))
    return resultado


def _limpiar_num(v: str) -> float | None:
    if not v or v in ("\xa0", "-", "S/M", ""):
        return None
    try:
        return float(v.replace(",", "").strip())
    except ValueError:
        return None


def _consultar_fecha(
    driver: webdriver.Chrome,
    fecha: date,
    primera_vez: bool,
    log_fn,
) -> list[dict]:
    if primera_vez:
        driver.get(URL_BASE)
        time.sleep(2)
        if not _esperar_pagina_lista(driver):
            log_fn("  [ERROR] La página no cargó. Verifica la conexión.")
            return []
        _esperar_telerik(driver)
        log_fn("  Página lista.")

    _set_fecha(driver, fecha)
    time.sleep(0.15)

    try:
        label_antes = driver.find_element(By.ID, ID_LBL_FECHA).text
    except NoSuchElementException:
        label_antes = ""

    try:
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, ID_BTN_CONSULTAR))
        )
        driver.execute_script("arguments[0].click();", btn)
    except (TimeoutException, NoSuchElementException):
        log_fn("  [ERROR] No se pudo hacer clic en Consultar.")
        return []

    _esperar_actualizacion(driver, label_antes)
    time.sleep(0.2)

    filas = _parsear_grid(driver)
    if not filas:
        log_fn("  Sin datos para esta fecha.")
    return filas


def _buscar_semillas(
    driver: webdriver.Chrome,
    df_raw: pd.DataFrame,
    fecha_inicio: date,
    log_fn,
) -> dict[str, dict]:
    """
    Para monedas con NaN al inicio del rango, busca su último valor real
    anterior a fecha_inicio (hasta _MAX_RETROCESO días atrás).
    Una sola consulta por fecha cubre todas las monedas pendientes.
    """
    semillas: dict[str, dict] = {}
    pendientes: set[str] = set()

    for moneda in df_raw["MONEDA"].unique():
        sub = df_raw[df_raw["MONEDA"] == moneda].sort_values("FECHA")
        primera = sub.iloc[0]
        if pd.isna(primera["Compra"]) or pd.isna(primera["Venta"]):
            pendientes.add(moneda)
            semillas[moneda] = {"Compra": None, "Venta": None}

    if not pendientes:
        return {}

    log_fn(f"  Buscando valor base para: {', '.join(sorted(pendientes))}...")

    fecha = fecha_inicio - timedelta(days=1)
    for _ in range(_MAX_RETROCESO):
        if not pendientes:
            break
        if fecha.weekday() >= 5:
            fecha -= timedelta(days=1)
            continue

        filas = _consultar_fecha(driver, fecha, False, lambda _: None)

        for fila in filas:
            moneda = _normalizar_moneda(fila.get("MONEDA", ""))
            if moneda not in pendientes:
                continue
            compra = _limpiar_num(fila.get("COMPRA (S/)", ""))
            venta  = _limpiar_num(fila.get("VENTA (S/)", ""))
            if semillas[moneda]["Compra"] is None and compra is not None:
                semillas[moneda]["Compra"] = compra
            if semillas[moneda]["Venta"] is None and venta is not None:
                semillas[moneda]["Venta"] = venta
            if semillas[moneda]["Compra"] is not None and semillas[moneda]["Venta"] is not None:
                pendientes.discard(moneda)

        if filas:
            cubiertas = sorted(m for m in semillas if m not in pendientes and semillas[m]["Compra"] is not None)
            if cubiertas:
                log_fn(f"  Valor base al {fecha.strftime('%d/%m/%Y')}: {', '.join(cubiertas)}")

        fecha -= timedelta(days=1)

    return semillas


def scrape_rango(
    fecha_inicio: date,
    fecha_fin: date,
    log_fn,
    cancelar_fn,
    progreso_fn=None,
) -> tuple[pd.DataFrame, dict]:
    acum        = []
    driver      = _crear_driver()
    primera_vez = True
    t0          = time.time()
    semillas    = {}

    # Feriados para todos los años del rango
    feriados: set[date] = set()
    for y in range(fecha_inicio.year, fecha_fin.year + 1):
        feriados.update(_feriados_peru(y))

    try:
        fechas = [fecha_inicio + timedelta(days=i)
                  for i in range((fecha_fin - fecha_inicio).days + 1)]

        # Días a consultar: hábiles y no feriados
        fechas_consulta   = [f for f in fechas if f.weekday() < 5 and f not in feriados]
        total             = len(fechas_consulta)
        procesados        = 0

        for fecha in fechas:
            if cancelar_fn():
                log_fn("Operación cancelada.")
                break
            if fecha.weekday() >= 5:
                continue
            if fecha in feriados:
                continue   # El ffill cubrirá este día

            procesados += 1
            log_fn(f"[{procesados}/{total}] {fecha.strftime('%d/%m/%Y')}")

            filas = _consultar_fecha(driver, fecha, primera_vez, log_fn)
            primera_vez = False

            for fila in filas:
                moneda = _normalizar_moneda(fila.get("MONEDA", ""))
                acum.append({
                    "FECHA":  fecha,
                    "CODIGO": _ISO.get(moneda, "???"),
                    "MONEDA": moneda,
                    "Compra": _limpiar_num(fila.get("COMPRA (S/)", "")),
                    "Venta":  _limpiar_num(fila.get("VENTA (S/)", "")),
                })

            if progreso_fn and total > 0:
                progreso_fn(int(procesados / total * 100))

        if acum and not cancelar_fn():
            df_temp  = pd.DataFrame(acum)
            semillas = _buscar_semillas(driver, df_temp, fecha_inicio, log_fn)

    finally:
        try:
            driver.quit()
        except Exception:
            pass

    minutos = max(1, round((time.time() - t0) / 60))
    log_fn(f"Tiempo total: ~{minutos} min.")

    if not acum:
        return pd.DataFrame(), {}
    return pd.DataFrame(acum).reset_index(drop=True), semillas


def aplicar_ffill(
    df: pd.DataFrame,
    fi: date,
    ff: date,
    semillas: dict | None = None,
) -> pd.DataFrame:
    """
    Reindexa a todos los días del rango y aplica ffill por moneda.
    Inyecta una fila semilla previa al rango para que el inicio no quede vacío.
    """
    if df.empty:
        return df

    todos_dias    = pd.date_range(fi, ff, freq="D")
    fecha_semilla = pd.Timestamp(fi) - timedelta(days=1)
    orden_moneda  = {m: i for i, m in enumerate(df["MONEDA"].unique())}
    partes        = []

    for moneda in df["MONEDA"].unique():
        sub = df[df["MONEDA"] == moneda].copy()
        sub = sub.set_index("FECHA")
        sub.index = pd.DatetimeIndex(sub.index)

        if semillas and moneda in semillas:
            seed = semillas[moneda]
            sub.loc[fecha_semilla] = {
                "CODIGO": _ISO.get(moneda, "???"),
                "MONEDA": moneda,
                "Compra": seed.get("Compra"),
                "Venta":  seed.get("Venta"),
            }

        rango_ext = pd.DatetimeIndex([fecha_semilla] + list(todos_dias))
        sub = sub.reindex(rango_ext.sort_values())
        sub["CODIGO"] = sub["CODIGO"].ffill().bfill()
        sub["MONEDA"] = sub["MONEDA"].ffill().bfill()
        sub["Compra"] = sub["Compra"].ffill().bfill()
        sub["Venta"]  = sub["Venta"].ffill().bfill()

        sub = sub.loc[todos_dias]
        sub.index.name = "FECHA"
        partes.append(sub.reset_index())

    out = pd.concat(partes, ignore_index=True)
    out["FECHA"] = out["FECHA"].dt.date
    out["_ord"]  = out["MONEDA"].map(orden_moneda)
    out = out.sort_values(["FECHA", "_ord"]).drop(columns="_ord").reset_index(drop=True)
    return out


# ============================================================
# EXPORTACION A EXCEL
# ============================================================

_COLOR_HDR = "1F4E79"
_COLOR_FNT = "FFFFFF"
_COLOR_ALT = "DDEBF7"


def _borde():
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def exportar_excel(df: pd.DataFrame, ruta: str) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="TipoCambio")

    cols       = list(df.columns)   # FECHA | CODIGO | MONEDA | Compra | Venta
    idx_fecha  = cols.index("FECHA")  + 1
    idx_compra = cols.index("Compra") + 1
    idx_venta  = cols.index("Venta")  + 1

    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font      = Font(bold=True, color=_COLOR_FNT, size=10)
        cell.fill      = PatternFill("solid", fgColor=_COLOR_HDR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _borde()
    ws.row_dimensions[1].height = 18

    for r, row in df.iterrows():
        er     = r + 2
        es_alt = r % 2 == 1
        for c, val in enumerate(row, 1):
            if not isinstance(val, (date, str)) and pd.isna(val):
                val = None
            cell = ws.cell(row=er, column=c, value=val)
            cell.border    = _borde()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if es_alt:
                cell.fill = PatternFill("solid", fgColor=_COLOR_ALT)
            if c == idx_fecha:
                cell.number_format = _FMT_DATE
                cell.alignment     = Alignment(horizontal="center", vertical="center")
            elif c in (idx_compra, idx_venta):
                cell.number_format = _FMT_NUM
                cell.alignment     = Alignment(horizontal="right", vertical="center")

    anchos = {"FECHA": 14, "CODIGO": 9, "MONEDA": 24, "Compra": 13, "Venta": 13}
    for c, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = anchos.get(col, 12)

    ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"
    tbl = Table(displayName=TABLA_EXCEL, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    wb.save(ruta)


# ============================================================
# INTERFAZ GRAFICA
# ============================================================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tipo de Cambio SBS - Extractor")
        self.resizable(False, False)
        self._cancelar = False
        self._hilo     = None
        self._construir_ui()
        self.protocol("WM_DELETE_WINDOW", self._cerrar_ventana)

    # ---- Construcción ----

    def _construir_ui(self):
        PAD = {"padx": 10, "pady": 6}

        # --- Rango de fechas ---
        frm_f = ttk.LabelFrame(self, text="Rango de fechas")
        frm_f.grid(row=0, column=0, sticky="ew", **PAD)
        frm_f.columnconfigure(1, weight=1)

        # Vars separadas para cada componente de fecha
        hoy = date.today()
        self._vi_d = tk.StringVar(value=f"{hoy.day:02d}")
        self._vi_m = tk.StringVar(value=f"{hoy.month:02d}")
        self._vi_y = tk.StringVar(value=str(hoy.year))
        self._vf_d = tk.StringVar(value=f"{hoy.day:02d}")
        self._vf_m = tk.StringVar(value=f"{hoy.month:02d}")
        self._vf_y = tk.StringVar(value=str(hoy.year))

        ttk.Label(frm_f, text="Fecha inicio:").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        ei_d, ei_m, ei_y = self._campos_fecha(frm_f, row=0, col=1,
                                               vd=self._vi_d, vm=self._vi_m, vy=self._vi_y)

        ttk.Label(frm_f, text="Fecha fin:").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        ef_d, ef_m, ef_y = self._campos_fecha(frm_f, row=1, col=1,
                                               vd=self._vf_d, vm=self._vf_m, vy=self._vf_y)

        # Auto-avance entre campos
        self._auto_avanzar(self._vi_d, 2, ei_m)
        self._auto_avanzar(self._vi_m, 2, ei_y)
        self._auto_avanzar(self._vi_y, 4, ef_d)
        self._auto_avanzar(self._vf_d, 2, ef_m)
        self._auto_avanzar(self._vf_m, 2, ef_y)

        self._var_info = tk.StringVar(value="")
        ttk.Label(frm_f, textvariable=self._var_info, foreground="#666666").grid(
            row=2, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 4))

        for v in (self._vi_d, self._vi_m, self._vi_y,
                  self._vf_d, self._vf_m, self._vf_y):
            v.trace_add("write", lambda *_: self._actualizar_info())
        self._actualizar_info()

        # --- Carpeta de salida ---
        frm_s = ttk.LabelFrame(self, text="Carpeta de salida")
        frm_s.grid(row=1, column=0, sticky="ew", **PAD)
        frm_s.columnconfigure(0, weight=1)

        self._var_carpeta = tk.StringVar(value=str(_directorio_exe()))
        self._ent_carpeta = ttk.Entry(frm_s, textvariable=self._var_carpeta)
        self._ent_carpeta.grid(row=0, column=0, sticky="ew", padx=(8, 2), pady=4)
        ttk.Button(frm_s, text="...", width=3,
                   command=self._elegir_carpeta).grid(row=0, column=1, padx=(0, 8), pady=4)

        self._var_prev = tk.StringVar(value="")
        ttk.Label(frm_s, textvariable=self._var_prev,
                  foreground="#2255AA", font=("Segoe UI", 8)).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 4))

        for v in (self._vi_d, self._vi_m, self._vi_y,
                  self._vf_d, self._vf_m, self._vf_y):
            v.trace_add("write", lambda *_: self._actualizar_prev())
        self._actualizar_prev()

        # --- Opciones ---
        frm_o = ttk.LabelFrame(self, text="Opciones")
        frm_o.grid(row=2, column=0, sticky="ew", **PAD)

        self._var_abrir = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_o, text="Abrir archivo al terminar",
                        variable=self._var_abrir).grid(
            row=0, column=0, sticky="w", padx=8, pady=4)

        self._var_abrir_carpeta = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm_o, text="Abrir carpeta al terminar",
                        variable=self._var_abrir_carpeta).grid(
            row=0, column=1, sticky="w", padx=8, pady=4)

        self._var_ffill = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_o, text="Incluir fines de semana y feriados",
                        variable=self._var_ffill).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))

        # --- Progreso ---
        frm_l = ttk.LabelFrame(self, text="Progreso")
        frm_l.grid(row=3, column=0, sticky="ew", **PAD)
        frm_l.columnconfigure(0, weight=1)

        self._var_estado = tk.StringVar(value="")
        ttk.Label(frm_l, textvariable=self._var_estado,
                  foreground="#1F4E79", font=("Segoe UI", 8)).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(4, 0))

        self._log = scrolledtext.ScrolledText(
            frm_l, height=11, width=70, state="disabled",
            font=("Consolas", 9), wrap="word",
        )
        self._log.grid(row=1, column=0, columnspan=2, padx=6, pady=(2, 2))

        ttk.Button(frm_l, text="Limpiar", width=8,
                   command=self._limpiar_log).grid(
            row=2, column=1, sticky="e", padx=8, pady=(0, 4))

        # --- Barra ---
        self._barra = ttk.Progressbar(self, mode="determinate", length=460, maximum=100)
        self._barra.grid(row=4, column=0, padx=10, pady=4, sticky="ew")

        # --- Botones ---
        frm_b = ttk.Frame(self)
        frm_b.grid(row=5, column=0, pady=8)
        self._btn_ej = ttk.Button(frm_b, text="Ejecutar", width=14, command=self._iniciar)
        self._btn_ej.pack(side="left", padx=8)
        self._btn_ca = ttk.Button(frm_b, text="Cancelar", width=14,
                                  command=self._solicitar_cancelar, state="disabled")
        self._btn_ca.pack(side="left", padx=8)

        self.columnconfigure(0, weight=1)

    def _campos_fecha(self, parent, row, col, vd, vm, vy):
        """Crea tres Entry (DD / MM / AAAA) con validación de dígitos."""
        vcmd2 = (self.register(lambda s: s == "" or (s.isdigit() and len(s) <= 2)), "%P")
        vcmd4 = (self.register(lambda s: s == "" or (s.isdigit() and len(s) <= 4)), "%P")

        frm = ttk.Frame(parent)
        frm.grid(row=row, column=col, sticky="w", padx=8, pady=4)

        ent_d = ttk.Entry(frm, textvariable=vd, width=3, validate="key", validatecommand=vcmd2)
        ent_d.pack(side="left")
        ttk.Label(frm, text=" / ").pack(side="left")
        ent_m = ttk.Entry(frm, textvariable=vm, width=3, validate="key", validatecommand=vcmd2)
        ent_m.pack(side="left")
        ttk.Label(frm, text=" / ").pack(side="left")
        ent_y = ttk.Entry(frm, textvariable=vy, width=5, validate="key", validatecommand=vcmd4)
        ent_y.pack(side="left")

        return ent_d, ent_m, ent_y

    def _auto_avanzar(self, var: tk.StringVar, max_len: int, siguiente: ttk.Entry):
        """Mueve el foco al siguiente campo cuando el actual está completo."""
        def _check(*_):
            v = var.get()
            if len(v) == max_len and v.isdigit():
                siguiente.focus_set()
                siguiente.select_range(0, tk.END)
        var.trace_add("write", _check)

    # ---- UI helpers ----

    def _pfecha(self, vd, vm, vy) -> date | None:
        try:
            return date(int(vy.get()), int(vm.get()), int(vd.get()))
        except (ValueError, OverflowError, TypeError):
            return None

    def _pfecha_con_error(self, vd, vm, vy, nombre: str) -> date | None:
        r = self._pfecha(vd, vm, vy)
        if r is None:
            messagebox.showerror(
                "Fecha inválida",
                f"La fecha {nombre} no es válida.\n"
                f"Verifica que día, mes y año sean correctos.",
            )
        return r

    def _actualizar_info(self):
        fi = self._pfecha(self._vi_d, self._vi_m, self._vi_y)
        ff = self._pfecha(self._vf_d, self._vf_m, self._vf_y)
        if fi and ff and fi <= ff:
            t = (ff - fi).days + 1
            self._var_info.set(f"{t} días en el rango")
        elif fi and ff and fi > ff:
            self._var_info.set("⚠ La fecha inicio es posterior a la fecha fin")
        else:
            self._var_info.set("")

    def _actualizar_prev(self):
        fi = self._pfecha(self._vi_d, self._vi_m, self._vi_y)
        ff = self._pfecha(self._vf_d, self._vf_m, self._vf_y)
        if fi and ff and fi <= ff:
            self._var_prev.set(f"Archivo: {_nombre_archivo(fi, ff)}")
        else:
            self._var_prev.set("")

    def _elegir_carpeta(self):
        c = filedialog.askdirectory(
            initialdir=self._var_carpeta.get() or str(_directorio_exe()),
            title="Seleccionar carpeta de salida",
        )
        if c:
            self._var_carpeta.set(c)

    def _log_msg(self, texto: str):
        def _w():
            self._log.configure(state="normal")
            self._log.insert("end", texto + "\n")
            self._log.see("end")
            self._log.configure(state="disabled")
        self.after(0, _w)

    def _set_estado(self, texto: str):
        self.after(0, lambda: self._var_estado.set(texto))

    def _limpiar_log(self):
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")

    def _set_progreso(self, v: int):
        self.after(0, lambda: self._barra.configure(value=v))

    def _set_controles(self, ejecutando: bool):
        ej = "disabled" if ejecutando else "normal"
        ca = "normal"   if ejecutando else "disabled"
        self._btn_ej.configure(state=ej)
        self._btn_ca.configure(state=ca)
        for w in (self._ent_carpeta,):
            w.configure(state="disabled" if ejecutando else "normal")
        if not ejecutando:
            self._barra.configure(value=0)
            self._set_estado("")

    def _solicitar_cancelar(self):
        self._cancelar = True
        self._log_msg(">>> Cancelando...")

    def _cerrar_ventana(self):
        if self._hilo and self._hilo.is_alive():
            if not messagebox.askyesno(
                "Proceso en curso",
                "Hay una consulta en ejecución. Cerrar ahora puede dejar "
                "Chrome abierto.\n¿Deseas cerrar de todas formas?",
            ):
                return
        self.destroy()

    def _log_con_estado(self, texto: str):
        self._log_msg(texto)
        if texto.startswith("["):
            self._set_estado(texto.strip())

    # ---- Ejecución ----

    def _iniciar(self):
        fi = self._pfecha_con_error(self._vi_d, self._vi_m, self._vi_y, "inicio")
        ff = self._pfecha_con_error(self._vf_d, self._vf_m, self._vf_y, "fin")
        if fi is None or ff is None:
            return
        if fi > ff:
            messagebox.showerror(
                "Rango inválido",
                f"La fecha de inicio ({fi.strftime('%d/%m/%Y')}) es posterior "
                f"a la fecha fin ({ff.strftime('%d/%m/%Y')}).",
            )
            return

        carpeta = self._var_carpeta.get().strip()
        if not carpeta:
            messagebox.showerror("Error", "Especifica la carpeta de salida.")
            return

        cp = Path(carpeta)
        if not cp.exists():
            if messagebox.askyesno("Carpeta no existe",
                                   f"La carpeta no existe:\n{carpeta}\n¿Deseas crearla?"):
                cp.mkdir(parents=True, exist_ok=True)
            else:
                return

        h    = _dias_habiles(fi, ff)
        est  = _fmt_duracion(h * _SEG_POR_DIA)
        ruta = str(cp / _nombre_archivo(fi, ff))

        self._cancelar = False
        self._set_controles(ejecutando=True)
        self._set_progreso(0)

        self._log_msg("=" * 55)
        self._log_msg(f"Inicio: {fi.strftime('%d/%m/%Y')}  |  Fin: {ff.strftime('%d/%m/%Y')}")
        self._log_msg(f"Duración estimada: {est}  |  Archivo: {Path(ruta).name}")
        self._log_msg("=" * 55)
        self._log_msg("Chrome se abrirá. No lo cierres durante el proceso.")

        self._hilo = threading.Thread(
            target=self._ejecutar_hilo,
            args=(fi, ff, ruta),
            daemon=True,
        )
        self._hilo.start()

    def _ejecutar_hilo(self, fi: date, ff: date, ruta: str):
        try:
            df, semillas = scrape_rango(
                fi, ff,
                log_fn=self._log_con_estado,
                cancelar_fn=lambda: self._cancelar,
                progreso_fn=self._set_progreso,
            )

            if self._cancelar:
                self._log_msg("Proceso cancelado. Archivo no generado.")
                return

            if df.empty:
                self._log_msg("[WARN] No se obtuvieron datos.")
                self.after(0, lambda: messagebox.showwarning(
                    "Sin datos",
                    "No se obtuvieron datos del sitio SBS.\n"
                    "Verifica que Chrome no esté siendo bloqueado.",
                ))
                return

            if self._var_ffill.get():
                self._log_msg("Expandiendo a todos los días del rango...")
                df = aplicar_ffill(df, fi, ff, semillas=semillas)

            self._log_msg("Generando archivo Excel...")
            exportar_excel(df, ruta)
            self._log_msg(f"Filas exportadas: {len(df)}")
            self._log_msg(f"Archivo: {ruta}")
            self._log_msg("Proceso completado.")
            self._set_progreso(100)
            self._set_estado("Completado.")

            if self._var_abrir.get():
                self.after(0, lambda: _abrir_archivo(ruta))
            if self._var_abrir_carpeta.get():
                self.after(0, lambda: _abrir_carpeta(str(Path(ruta).parent)))

        except Exception as exc:
            self._log_msg(f"[ERROR] {exc}")
            self.after(0, lambda: messagebox.showerror("Error", str(exc)))
        finally:
            self.after(0, lambda: self._set_controles(ejecutando=False))


# ============================================================
# UTILES DE SISTEMA
# ============================================================

def _abrir_archivo(ruta: str):
    try:
        if sys.platform == "win32":
            os.startfile(ruta)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", ruta])
        else:
            subprocess.Popen(["xdg-open", ruta])
    except Exception:
        pass


def _abrir_carpeta(ruta: str):
    try:
        if sys.platform == "win32":
            subprocess.Popen(["explorer", ruta])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", ruta])
        else:
            subprocess.Popen(["xdg-open", ruta])
    except Exception:
        pass


if __name__ == "__main__":
    app = App()
    app.mainloop()