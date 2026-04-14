"""
Scraper - Tipo de Cambio Ponderado SBS
Fuente: seriesH-tipo_cambio_moneda_excel.asp
Columnas de salida: FECHA | CODIGO | MONEDA | Compra | Venta
"""

import os
import sys
import time
import threading
import subprocess
from datetime import date, timedelta
from pathlib import Path
from urllib.parse import urlencode

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.scrolledtext as scrolledtext

import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# CONSTANTES
# ============================================================

URL_IFRAME    = (
    "https://www.sbs.gob.pe/app/stats/seriesH_TC-CV-Historico.asp"
)
URL_DESCARGA  = (
    "https://www.sbs.gob.pe/app/stats/seriesH-tipo_cambio_moneda_excel.asp"
)

TABLA_EXCEL   = "TipoCambioSBS"
_MAX_RETROCESO = 45

# Codigos conocidos del select del iframe (nombre SBS → codigo)
MONEDAS: dict[str, str] = {
    "02": "Dólar de N.A.",
    "11": "Dólar canadiense",
    "34": "Libra Esterlina",
    "38": "Yen japonés",
    "55": "Corona sueca",
    "57": "Franco suizo",
    "66": "Euro",
}

HEADERS_HTTP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-PE,es;q=0.9,en;q=0.8",
    "Referer": "https://www.sbs.gob.pe/",
}

_ISO: dict[str, str] = {
    "dólar de n.a.":          "USD",
    "dólar estadounidense":   "USD",
    "dólar australiano":      "AUD",
    "dólar canadiense":       "CAD",
    "dólar de hong kong":     "HKD",
    "dólar de singapur":      "SGD",
    "dólar neozelandés":      "NZD",
    "dólar de n.z.":          "NZD",
    "peso chileno":            "CLP",
    "peso colombiano":         "COP",
    "peso mexicano":           "MXN",
    "peso argentino":          "ARS",
    "real brasileño":          "BRL",
    "real brasilero":          "BRL",
    "libra esterlina":         "GBP",
    "yen japonés":             "JPY",
    "yuan chino":              "CNY",
    "renminbi":                "CNY",
    "won coreano":             "KRW",
    "franco suizo":            "CHF",
    "euro":                    "EUR",
    "corona sueca":            "SEK",
    "corona noruega":          "NOK",
    "corona danesa":           "DKK",
    "corona checa":            "CZK",
    "zloty polaco":            "PLN",
    "sol peruano":             "PEN",
    "nuevo sol":               "PEN",
}

_NOMBRE_USD = "Dólar Estadounidense"
_FMT_DATE   = "DD/MM/YYYY"
_FMT_NUM    = "#,##0.0000"


def _iso(nombre: str) -> str:
    return _ISO.get(nombre.strip().lower(), "???")


def _normalizar_moneda(nombre: str) -> str:
    return _NOMBRE_USD if nombre.strip().lower() == "dólar de n.a." else nombre.strip()


# ============================================================
# FERIADOS PERUANOS
# ============================================================

def _pascua(year: int) -> date:
    """Domingo de Pascua — algoritmo de Butcher."""
    a = year % 19; b = year // 100; c = year % 100
    d = b // 4;  e = b % 4;  f = (b + 8) // 25;  g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4;  k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day   = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def _feriados_peru(year: int) -> frozenset[date]:
    """Feriados nacionales peruanos: fijos + Semana Santa."""
    pascua = _pascua(year)
    fijos = [
        (1,  1), (5,  1), (6,  7), (6, 29),
        (7, 23), (7, 28), (7, 29), (8,  6),
        (8, 30), (10, 8), (11, 1), (12, 8),
        (12, 9), (12, 25),
    ]
    return frozenset(
        [date(year, m, d) for m, d in fijos]
        + [pascua - timedelta(days=3),
           pascua - timedelta(days=2)]
    )


# ============================================================
# UTILES
# ============================================================

def _directorio_exe() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _nombre_archivo(fi: date, ff: date) -> str:
    return f"TipoCambioPonderado_{fi.strftime('%d%m%Y')}_{ff.strftime('%d%m%Y')}.xlsx"


def _trimestre_anterior(hoy: date | None = None) -> tuple[date, date]:
    if hoy is None:
        hoy = date.today()
    q = (hoy.month - 1) // 3
    if q == 0:
        return date(hoy.year - 1, 10, 1), date(hoy.year - 1, 12, 31)
    starts = {1: (1, 1),  2: (4, 1),  3: (7, 1)}
    ends   = {1: (3, 31), 2: (6, 30), 3: (9, 30)}
    return (date(hoy.year, *starts[q]),
            date(hoy.year, *ends[q]))


def _construir_url(fecha1: str, fecha2: str, moneda: str, cierre: str = "") -> str:
    params = {"fecha1": fecha1, "fecha2": fecha2, "moneda": moneda, "cierre": cierre}
    return f"{URL_DESCARGA}?{urlencode(params, safe='/')}"


# ============================================================
# SESION INCAPSULA
# ============================================================

def _obtener_sesion(log_fn) -> requests.Session:
    """Abre Chrome visible, carga el iframe para resolver el challenge de
    Incapsula, transfiere las cookies a una sesion requests y cierra el browser.
    """
    log_fn("Abriendo Chrome para obtener sesión...")
    options = Options()
    options.add_argument("--window-size=1200,800")
    options.add_argument("--disable-extensions")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.page_load_strategy = "none"

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(15)
    try:
        driver.get(URL_IFRAME)
        time.sleep(6)
        cookies = driver.get_cookies()
    finally:
        driver.quit()

    session = requests.Session()
    for c in cookies:
        session.cookies.set(c["name"], c["value"], domain=c.get("domain", ""))

    log_fn(f"  Sesión lista ({len(cookies)} cookies transferidas).")
    return session


# ============================================================
# PARSEO DEL ENDPOINT
# ============================================================

def _limpiar_num(v: str) -> float | None:
    if not v or v.strip() in ("", "\xa0", "-", "S/M"):
        return None
    try:
        return float(v.replace(",", "").strip())
    except ValueError:
        return None


def _tiene_datos(resp: requests.Response) -> bool:
    content = resp.content
    if len(content) < 400:
        return False
    # Challenges de Incapsula siempre empiezan con <html.
    # Los datos reales empiezan con \r\n\r\n<TABLE BORDER=1>
    if content.lstrip(b"\r\n \t").lower().startswith(b"<html"):
        return False
    if b"FECHA" not in content:
        return False
    return True


def _parsear_respuesta(content: bytes) -> list[dict]:
    """Parsea el HTML de la respuesta y retorna lista de filas con
    FECHA (date), MONEDA (str), Compra (float|None), Venta (float|None).
    """
    try:
        html = content.decode("latin-1", errors="replace")
    except Exception:
        return []

    soup = BeautifulSoup(html, "html.parser")
    tabla = soup.find("table")
    if not tabla:
        return []

    filas = tabla.find_all("tr")
    if len(filas) < 2:
        return []

    resultado = []
    for fila in filas[1:]:
        celdas = [td.get_text(strip=True) for td in fila.find_all("td")]
        if len(celdas) < 4:
            continue
        fecha_str, moneda, compra_str, venta_str = (
            celdas[0].strip(),
            celdas[1].strip(),
            celdas[2].strip(),
            celdas[3].strip(),
        )
        try:
            d, m, y = fecha_str.split("/")
            fecha = date(int(y), int(m), int(d))
        except (ValueError, AttributeError):
            continue

        moneda = _normalizar_moneda(moneda)
        resultado.append({
            "FECHA":  fecha,
            "MONEDA": moneda,
            "Compra": _limpiar_num(compra_str),
            "Venta":  _limpiar_num(venta_str),
        })
    return resultado


# ============================================================
# SCRAPING PRINCIPAL
# ============================================================

def _descargar_moneda(
    session: requests.Session,
    codigo: str,
    nombre: str,
    fecha1_str: str,
    fecha2_str: str,
    log_fn,
) -> list[dict]:
    url = _construir_url(fecha1_str, fecha2_str, codigo)
    try:
        resp = session.get(url, headers=HEADERS_HTTP, timeout=120)
        resp.raise_for_status()
        if not _tiene_datos(resp):
            log_fn(f"  [{codigo}] {nombre}: sin datos en el rango.")
            return []
        filas = _parsear_respuesta(resp.content)
        log_fn(f"  [{codigo}] {nombre}: {len(filas)} registros.")
        return filas
    except requests.RequestException as exc:
        log_fn(f"  [{codigo}] {nombre}: ERROR — {exc}")
        return []


def _buscar_semillas(
    session: requests.Session,
    df_raw: pd.DataFrame,
    fecha_inicio: date,
    monedas_sel: dict[str, str],
    log_fn,
) -> dict[str, dict]:
    """Para monedas con Compra=NaN en la primera fila, busca el ultimo valor
    real previo a fecha_inicio descargando el rango historico de esa moneda.
    """
    semillas: dict[str, dict] = {}
    pendientes: list[tuple[str, str]] = []

    for codigo, nombre in monedas_sel.items():
        nombre_norm = _normalizar_moneda(nombre)
        sub = df_raw[df_raw["MONEDA"] == nombre_norm].sort_values("FECHA")
        if sub.empty or pd.isna(sub.iloc[0]["Compra"]):
            pendientes.append((codigo, nombre_norm))
            semillas[nombre_norm] = {"Compra": None, "Venta": None}

    if not pendientes:
        return {}

    log_fn(f"  Buscando valor base anterior para: {', '.join(n for _, n in pendientes)}...")

    # Buscar hasta 45 dias calendario antes de fecha_inicio
    fecha_seed_fin = fecha_inicio - timedelta(days=1)
    fecha_seed_ini = fecha_seed_fin - timedelta(days=_MAX_RETROCESO)
    fi_str = fecha_seed_ini.strftime("%d/%m/%Y")
    ff_str = fecha_seed_fin.strftime("%d/%m/%Y")

    for codigo, nombre_norm in pendientes:
        filas = _descargar_moneda(session, codigo, nombre_norm, fi_str, ff_str, lambda _: None)
        if not filas:
            continue
        # Tomar el ultimo registro no nulo
        for fila in reversed(filas):
            if semillas[nombre_norm]["Compra"] is None and fila["Compra"] is not None:
                semillas[nombre_norm]["Compra"] = fila["Compra"]
            if semillas[nombre_norm]["Venta"] is None and fila["Venta"] is not None:
                semillas[nombre_norm]["Venta"] = fila["Venta"]
            if semillas[nombre_norm]["Compra"] is not None:
                log_fn(f"  Valor base para {nombre_norm}: Compra={semillas[nombre_norm]['Compra']}")
                break

    return semillas


def scrape_rango(
    fecha_inicio: date,
    fecha_fin: date,
    monedas_sel: dict[str, str],
    log_fn,
    cancelar_fn,
    progreso_fn=None,
    estado_fn=None,
) -> tuple[pd.DataFrame, dict]:
    """Descarga el tipo de cambio del rango para cada moneda seleccionada.

    Retorna (DataFrame crudo, semillas para ffill).
    """
    session = _obtener_sesion(log_fn)

    fi_str = fecha_inicio.strftime("%d/%m/%Y")
    ff_str = fecha_fin.strftime("%d/%m/%Y")

    acum: list[dict] = []
    total = len(monedas_sel)

    log_fn(f"Descargando {total} moneda(s) [{fi_str} - {ff_str}]...")

    for i, (codigo, nombre) in enumerate(monedas_sel.items(), 1):
        if cancelar_fn():
            log_fn("Operación cancelada.")
            break

        if estado_fn:
            estado_fn(f"[{i}/{total}]  {nombre}")

        filas = _descargar_moneda(session, codigo, nombre, fi_str, ff_str, log_fn)
        for fila in filas:
            acum.append({
                "FECHA":  fila["FECHA"],
                "CODIGO": _iso(fila["MONEDA"]),
                "MONEDA": fila["MONEDA"],
                "Compra": fila["Compra"],
                "Venta":  fila["Venta"],
            })

        if progreso_fn:
            progreso_fn(int(i / total * 100))

    if not acum or cancelar_fn():
        return pd.DataFrame(), {}

    df = pd.DataFrame(acum).reset_index(drop=True)
    semillas = _buscar_semillas(session, df, fecha_inicio, monedas_sel, log_fn)
    return df, semillas


# ============================================================
# FFILL
# ============================================================

def aplicar_ffill(
    df: pd.DataFrame,
    fi: date,
    ff: date,
    semillas: dict | None = None,
) -> pd.DataFrame:
    """Reindexa a todos los dias del rango y aplica ffill por moneda.
    Inyecta fila semilla previa a fi para cubrir el bloque inicial sin datos.
    """
    if df.empty:
        return df

    todos_dias    = pd.date_range(fi, ff, freq="D")
    fecha_semilla = pd.Timestamp(fi) - timedelta(days=1)
    orden_moneda  = {m: i for i, m in enumerate(df["MONEDA"].unique())}
    partes        = []

    for moneda in df["MONEDA"].unique():
        sub = df[df["MONEDA"] == moneda].copy()
        sub = sub.set_index("FECHA")
        sub.index = pd.DatetimeIndex(sub.index)

        if semillas and moneda in semillas:
            seed = semillas[moneda]
            if seed.get("Compra") is not None:
                sub.loc[fecha_semilla] = {
                    "CODIGO": _iso(moneda),
                    "MONEDA": moneda,
                    "Compra": seed["Compra"],
                    "Venta":  seed.get("Venta"),
                }

        rango_ext = pd.DatetimeIndex([fecha_semilla] + list(todos_dias))
        sub = sub.reindex(rango_ext.sort_values())
        sub["CODIGO"] = sub["CODIGO"].ffill().bfill()
        sub["MONEDA"] = sub["MONEDA"].ffill().bfill()
        if sub["Compra"].notna().any():
            sub["Compra"] = sub["Compra"].ffill().bfill()
        if sub["Venta"].notna().any():
            sub["Venta"] = sub["Venta"].ffill().bfill()

        sub = sub.loc[todos_dias]
        sub.index.name = "FECHA"
        partes.append(sub.reset_index())

    out = pd.concat(partes, ignore_index=True)
    out["FECHA"] = out["FECHA"].dt.date
    out["_ord"]  = out["MONEDA"].map(orden_moneda)
    out = out.sort_values(["FECHA", "_ord"]).drop(columns="_ord").reset_index(drop=True)
    return out


# ============================================================
# EXPORTACION A EXCEL
# ============================================================

_COLOR_HDR = "1F4E79"
_COLOR_FNT = "FFFFFF"
_COLOR_ALT = "DDEBF7"


def _borde():
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def exportar_excel(df: pd.DataFrame, ruta: str) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="TipoCambio")

    cols       = list(df.columns)
    idx_fecha  = cols.index("FECHA")  + 1
    idx_compra = cols.index("Compra") + 1
    idx_venta  = cols.index("Venta")  + 1

    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font      = Font(bold=True, color=_COLOR_FNT, size=10)
        cell.fill      = PatternFill("solid", fgColor=_COLOR_HDR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _borde()
    ws.row_dimensions[1].height = 18

    for r, row in df.iterrows():
        er     = r + 2
        es_alt = r % 2 == 1
        for c, val in enumerate(row, 1):
            if not isinstance(val, (date, str)) and pd.isna(val):
                val = None
            cell = ws.cell(row=er, column=c, value=val)
            cell.border    = _borde()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if es_alt:
                cell.fill = PatternFill("solid", fgColor=_COLOR_ALT)
            if c == idx_fecha:
                cell.number_format = _FMT_DATE
                cell.alignment     = Alignment(horizontal="center", vertical="center")
            elif c in (idx_compra, idx_venta):
                cell.number_format = _FMT_NUM
                cell.alignment     = Alignment(horizontal="right", vertical="center")

    anchos = {"FECHA": 14, "CODIGO": 9, "MONEDA": 24, "Compra": 13, "Venta": 13}
    for c, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = anchos.get(col, 12)

    ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"
    tbl = Table(displayName=TABLA_EXCEL, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    wb.save(ruta)


# ============================================================
# INTERFAZ GRAFICA
# ============================================================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tipo de Cambio Ponderado SBS - Extractor")
        self.resizable(False, False)
        self._cancelar = False
        self._hilo     = None
        self._t_inicio = 0.0
        self._vars_moneda: dict[str, tk.BooleanVar] = {}
        self._construir_ui()
        self.protocol("WM_DELETE_WINDOW", self._cerrar_ventana)

    # ---- Construcción ----

    def _construir_ui(self):
        PAD = {"padx": 10, "pady": 6}
        fi_def, ff_def = _trimestre_anterior()

        # --- Rango de fechas ---
        frm_f = ttk.LabelFrame(self, text="Rango de fechas")
        frm_f.grid(row=0, column=0, sticky="ew", **PAD)

        self._vi_d = tk.StringVar(value=f"{fi_def.day:02d}")
        self._vi_m = tk.StringVar(value=f"{fi_def.month:02d}")
        self._vi_y = tk.StringVar(value=str(fi_def.year))
        self._vf_d = tk.StringVar(value=f"{ff_def.day:02d}")
        self._vf_m = tk.StringVar(value=f"{ff_def.month:02d}")
        self._vf_y = tk.StringVar(value=str(ff_def.year))

        ttk.Label(frm_f, text="Fecha inicio:").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        ei_d, ei_m, ei_y = self._campos_fecha(frm_f, 0, 1, self._vi_d, self._vi_m, self._vi_y)

        ttk.Label(frm_f, text="Fecha fin:").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        ef_d, ef_m, ef_y = self._campos_fecha(frm_f, 1, 1, self._vf_d, self._vf_m, self._vf_y)

        self._auto_avanzar(self._vi_d, 2, ei_m)
        self._auto_avanzar(self._vi_m, 2, ei_y)
        self._auto_avanzar(self._vi_y, 4, ef_d)
        self._auto_avanzar(self._vf_d, 2, ef_m)
        self._auto_avanzar(self._vf_m, 2, ef_y)

        self._var_info = tk.StringVar(value="")
        ttk.Label(frm_f, textvariable=self._var_info, foreground="#666666").grid(
            row=2, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 4))

        for v in (self._vi_d, self._vi_m, self._vi_y,
                  self._vf_d, self._vf_m, self._vf_y):
            v.trace_add("write", lambda *_: self._actualizar_info())
        self._actualizar_info()

        # --- Monedas ---
        frm_m = ttk.LabelFrame(self, text="Monedas")
        frm_m.grid(row=1, column=0, sticky="ew", **PAD)

        ttk.Label(frm_m, text="Selecciona las monedas a descargar:",
                  foreground="#444444").grid(
            row=0, column=0, columnspan=4, sticky="w", padx=8, pady=(4, 2))

        monedas_lista = sorted(MONEDAS.items(), key=lambda x: x[1])
        for idx, (codigo, nombre) in enumerate(monedas_lista):
            var = tk.BooleanVar(value=True)
            self._vars_moneda[codigo] = var
            col = idx % 2
            row = 1 + idx // 2
            ttk.Checkbutton(
                frm_m, text=nombre, variable=var
            ).grid(row=row, column=col, sticky="w", padx=12, pady=2)

        frm_btn_m = ttk.Frame(frm_m)
        frm_btn_m.grid(
            row=1 + (len(monedas_lista) + 1) // 2,
            column=0, columnspan=2, sticky="w", padx=8, pady=(2, 6),
        )
        ttk.Button(frm_btn_m, text="Todas",   width=8,
                   command=self._sel_todas).pack(side="left", padx=4)
        ttk.Button(frm_btn_m, text="Ninguna", width=8,
                   command=self._sel_ninguna).pack(side="left", padx=4)

        # --- Carpeta de salida ---
        frm_s = ttk.LabelFrame(self, text="Carpeta de salida")
        frm_s.grid(row=2, column=0, sticky="ew", **PAD)
        frm_s.columnconfigure(0, weight=1)

        self._var_carpeta = tk.StringVar(value=str(_directorio_exe()))
        self._ent_carpeta = ttk.Entry(frm_s, textvariable=self._var_carpeta)
        self._ent_carpeta.grid(row=0, column=0, sticky="ew", padx=(8, 2), pady=4)
        ttk.Button(frm_s, text="...", width=3,
                   command=self._elegir_carpeta).grid(row=0, column=1, padx=(0, 8), pady=4)

        self._var_prev = tk.StringVar(value="")
        ttk.Label(frm_s, textvariable=self._var_prev,
                  foreground="#2255AA", font=("Segoe UI", 8)).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 4))

        for v in (self._vi_d, self._vi_m, self._vi_y,
                  self._vf_d, self._vf_m, self._vf_y):
            v.trace_add("write", lambda *_: self._actualizar_prev())
        self._actualizar_prev()

        # --- Opciones ---
        frm_o = ttk.LabelFrame(self, text="Opciones")
        frm_o.grid(row=3, column=0, sticky="ew", **PAD)

        self._var_abrir = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_o, text="Abrir archivo al terminar",
                        variable=self._var_abrir).grid(
            row=0, column=0, sticky="w", padx=8, pady=4)

        self._var_abrir_carpeta = tk.BooleanVar(value=False)
        ttk.Checkbutton(frm_o, text="Abrir carpeta al terminar",
                        variable=self._var_abrir_carpeta).grid(
            row=0, column=1, sticky="w", padx=8, pady=4)

        self._var_ffill = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            frm_o,
            text="Llenar datos faltantes en fines de semana y feriados por el último valor disponible",
            variable=self._var_ffill,
        ).grid(row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(0, 6))

        # --- Progreso ---
        frm_l = ttk.LabelFrame(self, text="Progreso")
        frm_l.grid(row=4, column=0, sticky="ew", **PAD)
        frm_l.columnconfigure(0, weight=1)

        frm_st = ttk.Frame(frm_l)
        frm_st.grid(row=0, column=0, columnspan=2, sticky="ew", padx=6, pady=(4, 0))
        frm_st.columnconfigure(0, weight=1)

        self._var_estado = tk.StringVar(value="")
        ttk.Label(frm_st, textvariable=self._var_estado,
                  foreground="#1F4E79", font=("Segoe UI", 8)).grid(
            row=0, column=0, sticky="w")

        self._var_timer = tk.StringVar(value="")
        ttk.Label(frm_st, textvariable=self._var_timer,
                  foreground="#888888", font=("Consolas", 8)).grid(
            row=0, column=1, sticky="e")

        self._log = scrolledtext.ScrolledText(
            frm_l, height=11, width=70, state="disabled",
            font=("Consolas", 9), wrap="word",
        )
        self._log.grid(row=1, column=0, columnspan=2, padx=6, pady=(2, 2))

        ttk.Button(frm_l, text="Limpiar", width=8,
                   command=self._limpiar_log).grid(
            row=2, column=1, sticky="e", padx=8, pady=(0, 4))

        # --- Barra de progreso ---
        self._barra = ttk.Progressbar(self, mode="determinate", length=460, maximum=100)
        self._barra.grid(row=5, column=0, padx=10, pady=4, sticky="ew")

        # --- Botones ---
        frm_b = ttk.Frame(self)
        frm_b.grid(row=6, column=0, pady=8)
        self._btn_ej = ttk.Button(frm_b, text="Ejecutar", width=14, command=self._iniciar)
        self._btn_ej.pack(side="left", padx=8)
        self._btn_ca = ttk.Button(frm_b, text="Cancelar", width=14,
                                  command=self._solicitar_cancelar, state="disabled")
        self._btn_ca.pack(side="left", padx=8)

        self.columnconfigure(0, weight=1)

    def _campos_fecha(self, parent, row, col, vd, vm, vy):
        vcmd2 = (self.register(lambda s: s == "" or (s.isdigit() and len(s) <= 2)), "%P")
        vcmd4 = (self.register(lambda s: s == "" or (s.isdigit() and len(s) <= 4)), "%P")

        frm = ttk.Frame(parent)
        frm.grid(row=row, column=col, sticky="w", padx=8, pady=4)

        ent_d = ttk.Entry(frm, textvariable=vd, width=3, validate="key", validatecommand=vcmd2)
        ent_d.pack(side="left")
        ttk.Label(frm, text=" / ").pack(side="left")
        ent_m = ttk.Entry(frm, textvariable=vm, width=3, validate="key", validatecommand=vcmd2)
        ent_m.pack(side="left")
        ttk.Label(frm, text=" / ").pack(side="left")
        ent_y = ttk.Entry(frm, textvariable=vy, width=5, validate="key", validatecommand=vcmd4)
        ent_y.pack(side="left")

        return ent_d, ent_m, ent_y

    def _auto_avanzar(self, var: tk.StringVar, max_len: int, siguiente: ttk.Entry):
        def _check(*_):
            v = var.get()
            if len(v) == max_len and v.isdigit():
                siguiente.focus_set()
                siguiente.select_range(0, tk.END)
        var.trace_add("write", _check)

    # ---- UI helpers ----

    def _pfecha(self, vd, vm, vy) -> date | None:
        try:
            return date(int(vy.get()), int(vm.get()), int(vd.get()))
        except (ValueError, OverflowError):
            return None

    def _actualizar_info(self):
        fi = self._pfecha(self._vi_d, self._vi_m, self._vi_y)
        ff = self._pfecha(self._vf_d, self._vf_m, self._vf_y)
        if fi and ff:
            if fi > ff:
                self._var_info.set("⚠  La fecha inicio es posterior a la fecha fin")
            else:
                t = (ff - fi).days + 1
                self._var_info.set(f"{t} días en el rango")
        else:
            self._var_info.set("")

    def _actualizar_prev(self):
        fi = self._pfecha(self._vi_d, self._vi_m, self._vi_y)
        ff = self._pfecha(self._vf_d, self._vf_m, self._vf_y)
        if fi and ff and fi <= ff:
            self._var_prev.set(f"Archivo: {_nombre_archivo(fi, ff)}")
        else:
            self._var_prev.set("")

    def _sel_todas(self):
        for var in self._vars_moneda.values():
            var.set(True)

    def _sel_ninguna(self):
        for var in self._vars_moneda.values():
            var.set(False)

    def _elegir_carpeta(self):
        c = filedialog.askdirectory(
            initialdir=self._var_carpeta.get() or str(_directorio_exe()),
            title="Seleccionar carpeta de salida",
        )
        if c:
            self._var_carpeta.set(c)

    def _log_msg(self, texto: str):
        def _w():
            self._log.configure(state="normal")
            self._log.insert("end", texto + "\n")
            self._log.see("end")
            self._log.configure(state="disabled")
        self.after(0, _w)

    def _set_estado(self, texto: str):
        self.after(0, lambda: self._var_estado.set(texto))

    def _limpiar_log(self):
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")

    def _set_progreso(self, v: int):
        self.after(0, lambda: self._barra.configure(value=v))

    def _tick_timer(self):
        if self._hilo and self._hilo.is_alive():
            elapsed = int(time.time() - self._t_inicio)
            m, s = divmod(elapsed, 60)
            self._var_timer.set(f"{m:02d}:{s:02d}")
            self.after(1000, self._tick_timer)

    def _set_controles(self, ejecutando: bool):
        ej = "disabled" if ejecutando else "normal"
        ca = "normal"   if ejecutando else "disabled"
        self._btn_ej.configure(state=ej)
        self._btn_ca.configure(state=ca)
        self._ent_carpeta.configure(state="disabled" if ejecutando else "normal")
        for var_chk in self._vars_moneda.values():
            # Los checkbuttons no tienen estado editable directo;
            # se bloquean deshabilitando el frame (simplificado aqui)
            pass
        if not ejecutando:
            self._barra.configure(value=0)
            self._set_estado("")

    def _solicitar_cancelar(self):
        self._cancelar = True
        self._log_msg(">>> Cancelando...")

    def _cerrar_ventana(self):
        if self._hilo and self._hilo.is_alive():
            if not messagebox.askyesno(
                "Proceso en curso",
                "Hay una consulta en ejecución. Cerrar ahora puede dejar "
                "Chrome abierto.\n¿Deseas cerrar de todas formas?",
            ):
                return
        self.destroy()

    # ---- Ejecución ----

    def _iniciar(self):
        fi = self._pfecha(self._vi_d, self._vi_m, self._vi_y)
        ff = self._pfecha(self._vf_d, self._vf_m, self._vf_y)

        if fi is None:
            messagebox.showerror("Fecha inválida", "La fecha de inicio no es válida.")
            return
        if ff is None:
            messagebox.showerror("Fecha inválida", "La fecha fin no es válida.")
            return
        if fi > ff:
            messagebox.showerror(
                "Rango inválido",
                f"La fecha inicio ({fi.strftime('%d/%m/%Y')}) es posterior "
                f"a la fecha fin ({ff.strftime('%d/%m/%Y')}).",
            )
            return

        monedas_sel = {
            cod: MONEDAS[cod]
            for cod, var in self._vars_moneda.items()
            if var.get()
        }
        if not monedas_sel:
            messagebox.showerror("Sin monedas", "Selecciona al menos una moneda.")
            return

        carpeta = self._var_carpeta.get().strip()
        if not carpeta:
            messagebox.showerror("Error", "Especifica la carpeta de salida.")
            return

        cp = Path(carpeta)
        if not cp.exists():
            if messagebox.askyesno("Carpeta no existe",
                                   f"La carpeta no existe:\n{carpeta}\n¿Deseas crearla?"):
                cp.mkdir(parents=True, exist_ok=True)
            else:
                return

        ruta = str(cp / _nombre_archivo(fi, ff))

        if Path(ruta).exists():
            resp = messagebox.askyesnocancel(
                "Archivo ya existe",
                f"'{Path(ruta).name}' ya existe en la carpeta seleccionada.\n\n"
                "Sí       → Sobreescribir\n"
                "No       → Elegir otro nombre\n"
                "Cancelar → Cancelar la operación",
            )
            if resp is None:
                return
            if resp is False:
                nueva = filedialog.asksaveasfilename(
                    initialdir=str(cp),
                    initialfile=Path(ruta).name,
                    defaultextension=".xlsx",
                    filetypes=[("Excel", "*.xlsx")],
                    title="Guardar como",
                )
                if not nueva:
                    return
                ruta = nueva

        self._cancelar = False
        self._set_controles(ejecutando=True)
        self._set_progreso(0)
        self._var_timer.set("00:00")

        self._log_msg("=" * 55)
        self._log_msg(f"Inicio: {fi.strftime('%d/%m/%Y')}  |  Fin: {ff.strftime('%d/%m/%Y')}")
        self._log_msg(f"Monedas: {', '.join(monedas_sel.values())}")
        self._log_msg(f"Archivo: {Path(ruta).name}")
        self._log_msg("=" * 55)
        self._log_msg("Chrome se abrirá brevemente para autenticación.")

        self._t_inicio = time.time()
        self.after(1000, self._tick_timer)

        self._hilo = threading.Thread(
            target=self._ejecutar_hilo,
            args=(fi, ff, ruta, monedas_sel),
            daemon=True,
        )
        self._hilo.start()

    def _ejecutar_hilo(self, fi: date, ff: date, ruta: str, monedas_sel: dict[str, str]):
        try:
            df, semillas = scrape_rango(
                fi, ff,
                monedas_sel=monedas_sel,
                log_fn=self._log_msg,
                cancelar_fn=lambda: self._cancelar,
                progreso_fn=self._set_progreso,
                estado_fn=self._set_estado,
            )

            if self._cancelar:
                self._log_msg("Proceso cancelado. Archivo no generado.")
                return

            if df.empty:
                self._log_msg("[WARN] No se obtuvieron datos.")
                self.after(0, lambda: messagebox.showwarning(
                    "Sin datos",
                    "No se obtuvieron datos del sitio SBS.\n"
                    "Verifica la conexión o vuelve a intentarlo.",
                ))
                return

            if self._var_ffill.get():
                self._log_msg("Expandiendo a todos los días del rango...")
                df = aplicar_ffill(df, fi, ff, semillas=semillas)

            self._log_msg("Generando archivo Excel...")
            exportar_excel(df, ruta)
            self._log_msg(f"Filas exportadas: {len(df)}")
            self._log_msg(f"Archivo: {ruta}")
            self._log_msg("Proceso completado.")
            self._set_progreso(100)
            self._set_estado("Completado.")

            if self._var_abrir.get():
                self.after(0, lambda: _abrir_archivo(ruta))
            if self._var_abrir_carpeta.get():
                self.after(0, lambda: _abrir_carpeta(str(Path(ruta).parent)))

        except Exception as exc:
            self._log_msg(f"[ERROR] {exc}")
            self.after(0, lambda: messagebox.showerror("Error", str(exc)))
        finally:
            self.after(0, lambda: self._set_controles(ejecutando=False))


# ============================================================
# UTILES DE SISTEMA
# ============================================================

def _abrir_archivo(ruta: str):
    try:
        if sys.platform == "win32":
            os.startfile(ruta)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", ruta])
        else:
            subprocess.Popen(["xdg-open", ruta])
    except Exception:
        pass


def _abrir_carpeta(ruta: str):
    try:
        if sys.platform == "win32":
            subprocess.Popen(["explorer", ruta])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", ruta])
        else:
            subprocess.Popen(["xdg-open", ruta])
    except Exception:
        pass


if __name__ == "__main__":
    app = App()
    app.mainloop()