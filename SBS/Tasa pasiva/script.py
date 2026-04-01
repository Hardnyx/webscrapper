"""
Interfaz Tkinter para descargar las tasas pasivas de depósitos a plazo fijo
desde la SBS para un rango de fechas y exportarlas a Excel.
"""

import os
import re
import time
import threading
import traceback
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext

# Workaround Python 3.13: el GC destruye tk.Variable desde hilos secundarios
_orig_var_del = tk.Variable.__del__
def _safe_var_del(self):
    try:
        _orig_var_del(self)
    except RuntimeError:
        pass
tk.Variable.__del__ = _safe_var_del

# ==================== CONFIGURACIÓN BÁSICA ====================

URL = "https://www.sbs.gob.pe/app/pp/EstadisticasSAEEPortal/Paginas/TIPasivaDepositoEmpresa.aspx?tip=B"

INPUT_ID      = "ctl00_cphContent_rdpDate_dateInput"
BTN_ID        = "ctl00_cphContent_btnConsultar"
MAIN_TABLE_MN = "ctl00_cphContent_rpgActualPrimTablaMn_OT"
TAB_ME_ID     = "ctl00_cphContent_lbtnMex"
LBL_FECHA_ID  = "ctl00_cphContent_lblMensajeFecha"

WAIT_TIMEOUT = 20

RELEVANT_IDS_MN = {
    "ctl00_cphContent_rpgActualPrimTablaMn_OT",
    "ctl00_cphContent_rpgActualPrimTablaMn_ctl00_DataZone_DT",
    "ctl00_cphContent_rpgActualMn_OT",
    "ctl00_cphContent_rpgActualMn_ctl00_DataZone_DT",
}

RELEVANT_IDS_ME = {
    "ctl00_cphContent_rpgActualPrimTablaMex_OT",
    "ctl00_cphContent_rpgActualPrimTablaMex_ctl00_DataZone_DT",
    "ctl00_cphContent_rpgActualMex_OT",
    "ctl00_cphContent_rpgActualMex_ctl00_DataZone_DT",
}

RELEVANT_IDS = RELEVANT_IDS_MN | RELEVANT_IDS_ME

MAPPING_PARSE_MN = {
    "ctl00_cphContent_rpgActualPrimTablaMn_OT": "df_tabla10",
    "ctl00_cphContent_rpgActualMn_OT":          "df_tabla12",
}

MAPPING_PARSE_ME = {
    "ctl00_cphContent_rpgActualPrimTablaMex_OT": "df_tabla14",
    "ctl00_cphContent_rpgActualMex_OT":           "df_tabla16",
}

DESIRED_COLS = [
    "Banco",
    "Depósitos de Ahorro",
    "Hasta 30 días",
    "31-90 días",
    "91-180 días",
    "181-360 días",
    "Más de 360 días",
    "Depósitos a Plazo",
    "Depósitos CTS",
]

# ==================== FERIADOS 2025 ====================

HOLIDAYS_2025 = {
    datetime(2025,  1,  1).date(),
    datetime(2025,  4, 17).date(),
    datetime(2025,  4, 18).date(),
    datetime(2025,  5,  1).date(),
    datetime(2025,  6,  7).date(),
    datetime(2025,  6, 29).date(),
    datetime(2025,  7, 23).date(),
    datetime(2025,  7, 28).date(),
    datetime(2025,  7, 29).date(),
    datetime(2025,  8,  6).date(),
    datetime(2025,  8, 30).date(),
    datetime(2025, 10,  8).date(),
    datetime(2025, 11,  1).date(),
    datetime(2025, 12,  8).date(),
    datetime(2025, 12,  9).date(),
    datetime(2025, 12, 25).date(),
}


def is_holiday(date_obj: datetime) -> bool:
    try:
        d = date_obj.date()
    except AttributeError:
        d = date_obj
    return d in HOLIDAYS_2025 if d.year == 2025 else False

# ==================== UTILIDADES DE FECHA ====================


def parse_ddmmyyyy(s: str) -> datetime:
    return datetime.strptime(s, "%d/%m/%Y")


def date_range(start_str: str, end_str: str):
    start = parse_ddmmyyyy(start_str)
    end   = parse_ddmmyyyy(end_str)
    if end < start:
        raise ValueError("end_date debe ser >= start_date")
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)

# ==================== PARSE DE TABLAS HTML ====================


def build_header_with_spans(table):
    thead = table.find("thead")
    if not thead:
        return None
    trs = thead.find_all("tr")
    total_cols = 0
    for cell in trs[0].find_all(["th", "td"]):
        total_cols += int(cell.get("colspan", 1))
    nrows = len(trs)
    grid  = [["" for _ in range(total_cols)] for _ in range(nrows)]

    def first_empty_col(row_idx):
        for c in range(total_cols):
            if grid[row_idx][c] == "":
                return c
        return None

    for r_idx, tr in enumerate(trs):
        for cell in tr.find_all(["th", "td"]):
            text    = cell.get_text(strip=True)
            colspan = int(cell.get("colspan", 1))
            rowspan = int(cell.get("rowspan", 1))
            c       = first_empty_col(r_idx)
            if c is None:
                continue
            for cc in range(colspan):
                for rr in range(rowspan):
                    if r_idx + rr < nrows and c + cc < total_cols:
                        if grid[r_idx + rr][c + cc] == "":
                            grid[r_idx + rr][c + cc] = text

    final_names = []
    for col in range(total_cols):
        parts   = [grid[r][col] for r in range(nrows) if grid[r][col]]
        compact = []
        prev    = None
        for p in parts:
            if p != prev:
                compact.append(p)
            prev = p
        final_names.append(" - ".join(compact) if compact else f"col_{col}")
    return final_names


def extract_inner_data(table):
    tbody = table.find("tbody")
    if tbody:
        rows = tbody.find_all("tr")
    else:
        thead = table.find("thead")
        skip  = len(thead.find_all("tr")) if thead else 0
        rows  = table.find_all("tr")[skip:]
    data = []
    for tr in rows:
        tds   = tr.find_all("td")
        cells = [td.get_text("\n", strip=True) for td in tds]
        if any(c.strip() != "" for c in cells):
            data.append(cells)
    return data


def extract_banks(main_table):
    bank_cells = main_table.find_all("td", class_="rpgRowHeaderField")
    banks = [td.get_text(strip=True) for td in bank_cells if td.get_text(strip=True) != ""]
    if not banks:
        for tr in main_table.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) == 1:
                text = tds[0].get_text(strip=True)
                if text:
                    banks.append(text)
    return banks


def clean_num(x):
    if x is None:
        return float("nan")
    s = str(x).strip()
    if s in ("", "-", "--", "NA", "N/A"):
        return float("nan")
    s2 = s.replace(",", "").replace(" ", "")
    try:
        return float(s2)
    except Exception:
        return s


def ensure_banco_as_column_local(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    df = df.copy()
    if "Banco" in df.columns:
        cols = list(df.columns)
        if cols[0] != "Banco":
            cols.remove("Banco")
            cols.insert(0, "Banco")
            df = df[cols]
        return df
    try:
        idx_vals = list(df.index.astype(str))
        non_num  = sum(1 for v in idx_vals if not re.match(r"^[\d\.\- ]+$", v))
        if non_num / max(1, len(idx_vals)) > 0.5:
            df = df.reset_index()
    except Exception:
        pass
    if "Banco" not in df.columns and len(df.columns) > 0:
        df = df.rename(columns={df.columns[0]: "Banco"})
    if "Banco" in df.columns:
        cols = list(df.columns)
        cols.remove("Banco")
        cols.insert(0, "Banco")
        df = df[cols]
    return df


def detect_mapping(df: pd.DataFrame | None) -> dict:
    if df is None:
        return {k: None for k in DESIRED_COLS}
    df       = ensure_banco_as_column_local(df.copy())
    existing = list(df.columns)

    def find_col_by_keywords(keywords):
        for col in existing:
            low = col.lower()
            for kw in keywords:
                if kw.lower() in low:
                    return col
        return None

    mapping                        = {}
    mapping["Banco"]               = "Banco" if "Banco" in existing else None
    mapping["Depósitos de Ahorro"] = find_col_by_keywords(
        ["ahorro", "depósitos de ahorro", "depositos de ahorro"]
    )
    mapping["Hasta 30 días"]       = find_col_by_keywords(["hasta 30", "hasta 30 días", "0-30"])
    mapping["31-90 días"]          = find_col_by_keywords(["31-90", "31 - 90"])
    mapping["91-180 días"]         = find_col_by_keywords(["91-180", "91 - 180"])
    mapping["181-360 días"]        = find_col_by_keywords(["181-360", "181 - 360"])
    mapping["Más de 360 días"]     = find_col_by_keywords(["más de 360", "mas de 360"])
    # Columna total "Depósitos a Plazo": coincidencia exacta para no capturar
    # las subcolumnas "Depósitos a Plazo - Hasta 30 días", etc.
    mapping["Depósitos a Plazo"]   = next(
        (c for c in existing if c.lower() in ("depósitos a plazo", "depositos a plazo")),
        None,
    )
    mapping["Depósitos CTS"]       = find_col_by_keywords(["cts", "depósitos cts", "depositos cts"])

    used      = set(v for v in mapping.values() if v)
    remaining = [c for c in existing if c not in used]
    for d in DESIRED_COLS:
        if d == "Banco":
            continue
        if not mapping.get(d):
            mapping[d] = remaining.pop(0) if remaining else None
    return mapping


def harmonize_to_desired(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    df       = ensure_banco_as_column_local(df.copy())
    existing = list(df.columns)
    if existing == DESIRED_COLS:
        return df[DESIRED_COLS].copy()
    mapping  = detect_mapping(df)
    n        = len(df)
    new_rows = {}
    for d in DESIRED_COLS:
        src         = mapping.get(d)
        new_rows[d] = df[src].values if (src and src in df.columns) else np.array([np.nan] * n)
    new_df = pd.DataFrame(new_rows)
    if "Banco" in new_df.columns:
        new_df["Banco"] = new_df["Banco"].astype(str).replace("nan", "")
    return new_df


def html_has_data(html_text: str, relevant_ids=None):
    soup   = BeautifulSoup(html_text, "html.parser")
    tables = soup.find_all("table")
    ids    = {t.get("id") for t in tables if t.get("id")}
    if relevant_ids is None:
        relevant_ids = RELEVANT_IDS
    return len(ids & relevant_ids) > 0, ids


def parse_tables_from_html(html_text: str, mapping: dict[str, str]):
    soup_local = BeautifulSoup(html_text, "html.parser")
    created:    dict[str, pd.DataFrame] = {}
    debug_info = []

    def align_header_and_rows(col_names, rows_clean):
        if not col_names:
            return col_names, rows_clean
        n_header = len(col_names)
        max_len  = max((len(r) for r in rows_clean), default=0)

        def looks_like_entity_header(h):
            h0 = (h or "").strip().lower()
            return (not h0) or ("empresa" in h0) or ("entidad" in h0) or ("sistema" in h0)

        if max_len == n_header - 1:
            if looks_like_entity_header(col_names[0]):
                col_names = col_names[1:]
                n_header -= 1
        elif max_len == n_header + 1:
            col_names  = col_names[1:]
            rows_clean = [r[1:] for r in rows_clean]
            n_header  -= 1

        fixed_rows = []
        for r in rows_clean:
            r2 = list(r[:n_header])
            if len(r2) < n_header:
                r2 += [""] * (n_header - len(r2))
            fixed_rows.append(r2)
        return col_names, fixed_rows

    for tabla_id, base_var in mapping.items():
        main_tbl = soup_local.find("table", {"id": tabla_id})
        if main_tbl is None:
            continue

        candidates = main_tbl.find_all("table")
        if candidates:
            inner_tbl = max(
                candidates,
                key=lambda t: sum(len(tr.find_all("td")) for tr in t.find_all("tr")),
            )
        else:
            inner_tbl = main_tbl

        col_names = build_header_with_spans(inner_tbl)
        rows      = extract_inner_data(inner_tbl)
        banks     = extract_banks(main_tbl)

        if not col_names:
            maxcols   = max((len(r) for r in rows), default=0)
            col_names = [f"col_{j}" for j in range(maxcols)]

        rows_clean = []
        for r in rows:
            r_clean = [cell.split("\n")[0].strip() if isinstance(cell, str) else cell for cell in r]
            rows_clean.append(r_clean)

        col_names, rows_clean = align_header_and_rows(col_names, rows_clean)

        if not rows_clean:
            df_inner = pd.DataFrame(columns=["Banco"] + col_names)
        else:
            df_inner = pd.DataFrame(rows_clean, columns=col_names)
            if banks and len(banks) == len(df_inner):
                df_inner.insert(0, "Banco", banks)
            else:
                df_inner.insert(0, "Banco", [""] * len(df_inner))

        for col in df_inner.columns:
            if col != "Banco":
                df_inner[col] = df_inner[col].apply(clean_num)

        df_inner = df_inner.reset_index(drop=True)
        created[base_var] = df_inner
        debug_info.append((base_var, df_inner.shape))

    return created, debug_info


def split_person_tables(df_persona: pd.DataFrame | None):
    if df_persona is None or not isinstance(df_persona, pd.DataFrame) or df_persona.empty:
        return None, None

    dfp      = ensure_banco_as_column_local(df_persona.copy())
    cols     = [c for c in dfp.columns if c != "Banco"]
    nat_cols = [c for c in cols if re.search(r"Natur", c, re.I)]
    jur_cols = [c for c in cols if re.search(r"Jur", c, re.I) or re.search(r"Jurid", c, re.I)]

    if not nat_cols and not jur_cols and cols:
        half     = len(cols) // 2
        nat_cols = cols[:half]
        jur_cols = cols[half:]

    def build_person_subdf(subcols):
        if not subcols:
            return None
        df2      = dfp[["Banco"] + subcols].copy()
        newnames = {}
        for c in subcols:
            if " - " in c:
                newnames[c] = c.split(" - ", 1)[1].strip()
            else:
                m = re.search(r"(Hasta.*|31-90.*|91-180.*|181-360.*|Más de.*)", c, re.I)
                newnames[c] = m.group(1).strip() if m else c
        df2 = df2.rename(columns=newnames)
        return ensure_banco_as_column_local(df2)

    return build_person_subdf(nat_cols), build_person_subdf(jur_cols)

# ==================== ARMADO DE FILAS PARA EXPORTACIÓN ====================


def build_general_rows_for_date(date_str: str, mn_general, me_general) -> pd.DataFrame | None:
    rows = []
    for df_src, moneda in [(mn_general, "Moneda Nacional"), (me_general, "Moneda Extranjera")]:
        if isinstance(df_src, pd.DataFrame):
            dh = harmonize_to_desired(df_src)
            if dh is not None and not dh.empty:
                df  = dh.copy()
                df.insert(0, "Tipo de Moneda", moneda)
                df.insert(0, "Fecha", date_str)
                cols = ["Fecha", "Tipo de Moneda", "Banco"] + [c for c in DESIRED_COLS if c != "Banco"]
                for c in cols:
                    if c not in df.columns:
                        df[c] = np.nan
                rows.append(df[cols])
    return pd.concat(rows, ignore_index=True) if rows else None


def build_person_rows_for_date(
    date_str: str, mn_nat, mn_jur, me_nat, me_jur
) -> pd.DataFrame | None:
    out_cols = [
        "Fecha", "Tipo de Moneda", "Tipo Persona", "Banco",
        "Hasta 30 días", "31-90 días", "91-180 días", "181-360 días", "Más de 360 días",
    ]
    rows = []

    def prepare_person_df(df, tipo_moneda, tipo_persona):
        if df is None or not isinstance(df, pd.DataFrame) or df.empty:
            return None
        df2     = ensure_banco_as_column_local(df.copy())
        mapping = detect_mapping(df2)
        minimal = pd.DataFrame()
        minimal["Banco"] = df2["Banco"].astype(str)
        for c in ["Hasta 30 días", "31-90 días", "91-180 días", "181-360 días", "Más de 360 días"]:
            src         = mapping.get(c)
            minimal[c]  = df2[src].values if (src and src in df2.columns) else np.nan
        minimal.insert(0, "Tipo Persona",   tipo_persona)
        minimal.insert(0, "Tipo de Moneda", tipo_moneda)
        minimal.insert(0, "Fecha",          date_str)
        return minimal[out_cols]

    for dfp, moneda, persona in [
        (mn_nat, "Moneda Nacional",   "Natural"),
        (mn_jur, "Moneda Nacional",   "Jurídica"),
        (me_nat, "Moneda Extranjera", "Natural"),
        (me_jur, "Moneda Extranjera", "Jurídica"),
    ]:
        p = prepare_person_df(dfp, moneda, persona)
        if p is not None:
            rows.append(p)
    return pd.concat(rows, ignore_index=True) if rows else None

# ==================== UPSERT A EXCEL ====================


def upsert_to_excel_accum(path: str, new_df: pd.DataFrame, key_cols, sheet_name="Datos"):
    if new_df is None or new_df.empty:
        return "no_rows"

    if os.path.exists(path):
        try:
            existing = pd.read_excel(path, sheet_name=sheet_name)
        except Exception:
            try:
                existing = pd.read_excel(path)
            except Exception:
                existing = pd.DataFrame()
    else:
        existing = pd.DataFrame()

    def normalize_keys(df_keys: pd.DataFrame) -> pd.DataFrame:
        dfk = df_keys.copy()
        if "Fecha" in dfk.columns:
            dfk["Fecha"] = pd.to_datetime(dfk["Fecha"], dayfirst=True, errors="coerce").dt.date
        for c in dfk.columns:
            dfk[c] = dfk[c].astype(str)
        return dfk

    if existing.empty:
        merged = new_df.copy()
    else:
        if all(k in new_df.columns for k in key_cols) and all(k in existing.columns for k in key_cols):
            existing_keys     = normalize_keys(existing[key_cols])
            new_keys          = normalize_keys(new_df[key_cols])
            existing_key_tups = existing_keys.apply(tuple, axis=1)
            new_key_tups      = set(new_keys.apply(tuple, axis=1))
            keep_mask         = ~existing_key_tups.isin(new_key_tups)
            existing_kept     = existing[keep_mask].copy()
            merged            = pd.concat([existing_kept, new_df], ignore_index=True, sort=False)
        else:
            merged = pd.concat([existing, new_df], ignore_index=True, sort=False)

    if "Fecha" in merged.columns:
        merged["Fecha"] = pd.to_datetime(merged["Fecha"], dayfirst=True, errors="coerce")
        try:
            merged = merged.sort_values(["Fecha"]).reset_index(drop=True)
        except Exception:
            pass

    try:
        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
            merged.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets.get(sheet_name)
            if ws is None:
                return "written_no_ws"

            nrows = merged.shape[0] + 1
            ncols = merged.shape[1]

            if "Fecha" in merged.columns:
                try:
                    fecha_col_idx = merged.columns.get_loc("Fecha") + 1
                    for r in range(2, nrows + 1):
                        cell = ws.cell(row=r, column=fecha_col_idx)
                        if isinstance(cell.value, str):
                            try:
                                dt = pd.to_datetime(cell.value, dayfirst=True, errors="coerce")
                                if not pd.isna(dt):
                                    cell.value = dt.to_pydatetime()
                            except Exception:
                                pass
                        cell.number_format = "DD/MM/YYYY"
                except Exception:
                    pass

            for col_idx in range(1, ncols + 1):
                col_letter = get_column_letter(col_idx)
                header     = str(merged.columns[col_idx - 1])
                max_len    = len(header)
                try:
                    col_values  = merged.iloc[:, col_idx - 1].astype(str).fillna("")
                    max_val_len = col_values.map(len).max() if not col_values.empty else 0
                    max_len     = max(max_len, int(max_val_len))
                except Exception:
                    pass
                try:
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
                except Exception:
                    pass

            try:
                last_col             = get_column_letter(ncols)
                ref                  = f"A1:{last_col}{nrows}"
                base_table_name      = (
                    "TasaPasivaTipoPersona" if "Tipo Persona" in merged.columns
                    else "TasaPasiva"
                )
                existing_table_names = set(ws.tables.keys())
                table_name_candidate = base_table_name
                i = 1
                while table_name_candidate in existing_table_names:
                    table_name_candidate = f"{base_table_name}_{i}"
                    i += 1
                table = Table(displayName=table_name_candidate, ref=ref)
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False, showLastColumn=False,
                    showRowStripes=True,   showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)
            except Exception:
                pass

        return "written"
    except Exception as e:
        return f"error_write: {e}"

# ==================== HELPERS SELENIUM ====================

DATE_RE = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")


def get_effective_date_from_label(driver):
    try:
        txt = driver.find_element(By.ID, LBL_FECHA_ID).text.strip()
    except Exception:
        return None, None
    matches = DATE_RE.findall(txt)
    return (matches[-1] if matches else None), txt


def set_sbs_date(driver, date_str: str):
    """
    Establece la fecha via API JS de Telerik para actualizar el ClientState
    que el servidor lee en el postback.
    """
    wait     = WebDriverWait(driver, WAIT_TIMEOUT)
    attempts = 3
    for i in range(attempts):
        try:
            driver.execute_script(
                """
                var rdp = $find('ctl00_cphContent_rdpDate');
                if (rdp) {
                    var parts = arguments[0].split('/');
                    var d = new Date(parseInt(parts[2]), parseInt(parts[1]) - 1, parseInt(parts[0]));
                    rdp.set_selectedDate(d);
                }
                """,
                date_str,
            )
            time.sleep(0.3)
            inp = driver.find_element(By.ID, INPUT_ID)
            if (inp.get_attribute("value") or "").strip() != date_str:
                driver.execute_script(
                    """
                    var el = arguments[0];
                    var setter = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype, 'value').set;
                    setter.call(el, arguments[1]);
                    el.dispatchEvent(new Event('input',  { bubbles: true }));
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                    el.blur();
                    """,
                    inp, date_str,
                )
                time.sleep(0.3)
            return
        except StaleElementReferenceException:
            if i == attempts - 1:
                raise
            time.sleep(0.5)
        except Exception:
            try:
                inp = wait.until(EC.element_to_be_clickable((By.ID, INPUT_ID)))
                driver.execute_script(
                    """
                    var el = arguments[0];
                    el.value = arguments[1];
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                    el.blur();
                    """,
                    inp, date_str,
                )
                time.sleep(0.4)
                return
            except Exception:
                if i == attempts - 1:
                    raise
                time.sleep(0.5)


def click_consultar_and_wait(driver, date_str: str):
    """
    Clic en Consultar y espera a que el label refleje la fecha solicitada
    en contexto de Moneda Nacional (postback MN completado).
    """
    wait     = WebDriverWait(driver, WAIT_TIMEOUT)
    attempts = 3
    old_eff, old_txt = get_effective_date_from_label(driver)

    for i in range(attempts):
        try:
            btn = wait.until(EC.element_to_be_clickable((By.ID, BTN_ID)))
            driver.execute_script("arguments[0].scrollIntoView(true);", btn)
            time.sleep(0.2)
            btn.click()

            def _mn_ready(drv):
                eff, txt = get_effective_date_from_label(drv)
                if eff is None:
                    return False
                return txt != old_txt or eff != old_eff

            try:
                wait.until(_mn_ready)
            except Exception:
                pass

            wait.until(EC.presence_of_element_located((By.ID, MAIN_TABLE_MN)))
            return get_effective_date_from_label(driver)

        except StaleElementReferenceException:
            if i == attempts - 1:
                raise
            time.sleep(0.5)
        except Exception:
            try:
                driver.execute_script(
                    "document.getElementById(arguments[0]).click();", BTN_ID
                )
                wait.until(EC.presence_of_element_located((By.ID, MAIN_TABLE_MN)))
                return get_effective_date_from_label(driver)
            except Exception:
                if i == attempts - 1:
                    raise
                time.sleep(0.5)

    return None, None


def click_me_tab_and_wait(driver, date_str: str) -> bool:
    """
    Clic en el tab Moneda Extranjera y espera a que el postback AJAX termine.

    Estrategia:
    1. Verificar que no hay postback en curso (isInAsyncPostBack = False).
    2. Verificar que el label dice "Extranjera... al date_str".
    Si el timeout expira sin confirmación, retorna False para que el caller
    no capture datos ME stale.
    """
    wait     = WebDriverWait(driver, WAIT_TIMEOUT)
    attempts = 3

    for i in range(attempts):
        try:
            mex_tab = wait.until(EC.element_to_be_clickable((By.ID, TAB_ME_ID)))
            mex_tab.click()

            def _me_ready(drv):
                # Primero verificar que el postback AJAX terminó
                try:
                    is_posting = drv.execute_script(
                        "try {"
                        "  var prm = Sys.WebForms.PageRequestManager.getInstance();"
                        "  return prm.get_isInAsyncPostBack();"
                        "} catch(e) { return false; }"
                    )
                    if is_posting:
                        return False
                except Exception:
                    pass
                # Luego verificar que el label refleja ME + fecha correcta
                eff, txt = get_effective_date_from_label(drv)
                if not txt or eff is None:
                    return False
                return "Extranjera" in txt and eff == date_str

            try:
                wait.until(_me_ready)
                time.sleep(0.3)
                return True
            except Exception:
                # El wait expiró: postback no confirmado, NO retornar True
                return False

        except StaleElementReferenceException:
            if i == attempts - 1:
                return False
            time.sleep(0.5)
        except Exception:
            if i == attempts - 1:
                return False
            time.sleep(0.5)
    return False

# ==================== PROCESO PRINCIPAL ====================


def run_date_range(
    start_date_str:    str,
    end_date_str:      str,
    out_base_dir:      str | None = None,
    log_fn             = print,
    simple_log_fn      = None,
    import_natural:    bool = True,
    import_juridica:   bool = True,
    import_general:    bool = True,
    remove_html_after: bool = False,
    skip_weekends:     bool = True,
):
    if simple_log_fn is not None:
        simple_log_fn(f"Iniciando proceso desde {start_date_str} hasta {end_date_str}.")

    range_tag = f"{start_date_str.replace('/', '')}-{end_date_str.replace('/', '')}"
    base_dir  = out_base_dir or os.getcwd()
    out_dir   = os.path.join(base_dir, f"TPF_{range_tag}")
    os.makedirs(out_dir, exist_ok=True)

    file_general   = os.path.join(out_dir, f"TPF_General_{range_tag}.xlsx")
    file_by_person = os.path.join(out_dir, f"TPF_Persona_{range_tag}.xlsx")

    log_fn(f"[INFO] Carpeta de salida: {out_dir}")

    driver = webdriver.Chrome()
    driver.get(URL)
    time.sleep(2)

    overall_export_results = []

    try:
        for dt in date_range(start_date_str, end_date_str):
            date_str   = dt.strftime("%d/%m/%Y")
            is_weekend = dt.weekday() >= 5

            if skip_weekends and is_weekend:
                log_fn(f"[INFO] Saltando fin de semana {date_str}.")
                if simple_log_fn is not None:
                    simple_log_fn(f"{date_str}: omitido (fin de semana).")
                continue

            if is_holiday(dt):
                log_fn(f"[INFO] Saltando feriado {date_str}.")
                if simple_log_fn is not None:
                    simple_log_fn(f"{date_str}: omitido (feriado).")
                continue

            date_suffix       = dt.strftime("%d%m%Y")
            had_data_for_date = False

            log_fn("")
            log_fn("=" * 60)
            log_fn(f"Procesando fecha {date_str} (sufijo: {date_suffix})")
            log_fn("=" * 60)

            html_path = os.path.join(out_dir, f"sbs_fuente_{date_suffix}.html")

            # ===== Paso 1: Click Consultar → MN se actualiza =====
            try:
                set_sbs_date(driver, date_str)
                eff_date, lbl_txt = click_consultar_and_wait(driver, date_str)
            except Exception as e:
                log_fn(f"[ERROR] No se pudo consultar la fecha {date_str}: {e}")
                if simple_log_fn is not None:
                    simple_log_fn(f"{date_str}: error al consultar en la SBS.")
                continue

            if eff_date and eff_date != date_str:
                log_fn(
                    f"[INFO] {date_str}: SBS reporta data vigente al {eff_date}. "
                    f"Se omite para evitar fecha inválida."
                )
                if simple_log_fn is not None:
                    simple_log_fn(f"{date_str}: omitido (SBS vigente al {eff_date}).")
                continue

            # Capturar MN desde el page_source post-Consultar
            html_mn = driver.page_source

            # ===== Paso 2: Click tab ME → ME se actualiza =====
            # El label cambiará a "Extranjera... al DATE" cuando el postback ME termine.
            ok_me = click_me_tab_and_wait(driver, date_str)
            if not ok_me:
                log_fn(f"[WARN] {date_str}: postback ME no confirmado. Datos ME omitidos para esta fecha.")

            # Capturar ME solo si el postback fue confirmado
            html_me = driver.page_source if ok_me else None

            # Guardar HTML del día: si ME fue confirmado usamos ese, si no usamos MN
            html_for_file = html_me if html_me is not None else html_mn
            try:
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(html_for_file)
            except Exception:
                pass

            # ===== Parsear MN desde html_mn (post-Consultar) =====
            mn_general = mn_nat = mn_jur = None
            has_mn, ids_mn = html_has_data(html_mn, RELEVANT_IDS_MN)
            if not has_mn:
                log_fn(f"[INFO] No hay tablas MN para {date_str}.")
            else:
                log_fn(f"[INFO] Tablas MN={len(ids_mn & RELEVANT_IDS_MN)}  detectadas")
                mn_tables, created_mn = parse_tables_from_html(html_mn, MAPPING_PARSE_MN)
                for base_name, shape in created_mn:
                    log_fn(f"  MN: {base_name} shape={shape}")
                mn_general = mn_tables.get("df_tabla10")
                mn_nat, mn_jur = split_person_tables(mn_tables.get("df_tabla12"))

            # ===== Parsear ME desde html_me (post-tab-click confirmado) =====
            me_general = me_nat = me_jur = None
            if html_me is None:
                log_fn(f"[INFO] {date_str}: ME omitido (postback no confirmado).")
            else:
                has_me, ids_me = html_has_data(html_me, RELEVANT_IDS_ME)
                if not has_me:
                    log_fn(f"[INFO] No hay tablas ME para {date_str}.")
                else:
                    log_fn(f"[INFO] Tablas ME={len(ids_me & RELEVANT_IDS_ME)}  detectadas")
                    me_tables, created_me = parse_tables_from_html(html_me, MAPPING_PARSE_ME)
                    for base_name, shape in created_me:
                        log_fn(f"  ME: {base_name} shape={shape}")
                    me_general = me_tables.get("df_tabla14")
                    me_nat, me_jur = split_person_tables(me_tables.get("df_tabla16"))

            # ===== Armar y exportar =====
            general_rows = (
                build_general_rows_for_date(date_str, mn_general, me_general)
                if import_general else None
            )
            person_rows = build_person_rows_for_date(date_str, mn_nat, mn_jur, me_nat, me_jur)

            if person_rows is not None:
                if not import_natural:
                    person_rows = person_rows[
                        person_rows["Tipo Persona"] != "Natural"
                    ].reset_index(drop=True)
                if not import_juridica:
                    person_rows = person_rows[
                        person_rows["Tipo Persona"] != "Jurídica"
                    ].reset_index(drop=True)
                if person_rows.empty:
                    person_rows = None

            if general_rows is not None and not general_rows.empty:
                had_data_for_date = True
                res = upsert_to_excel_accum(
                    file_general, general_rows, ["Fecha", "Tipo de Moneda", "Banco"]
                )
                overall_export_results.append(
                    (file_general, date_suffix, res, f"rows={len(general_rows)}")
                )
                log_fn(
                    f"[EXPORT] General -> {os.path.basename(file_general)}: "
                    f"{res} ({len(general_rows)} filas)"
                )

            if person_rows is not None and not person_rows.empty:
                had_data_for_date = True
                res2 = upsert_to_excel_accum(
                    file_by_person, person_rows,
                    ["Fecha", "Tipo de Moneda", "Tipo Persona", "Banco"],
                )
                overall_export_results.append(
                    (file_by_person, date_suffix, res2, f"rows={len(person_rows)}")
                )
                log_fn(
                    f"[EXPORT] Por tipo de persona -> {os.path.basename(file_by_person)}: "
                    f"{res2} ({len(person_rows)} filas)"
                )

            if simple_log_fn is not None:
                if had_data_for_date:
                    simple_log_fn(f"{date_str} descargado exitosamente ✅")
                else:
                    simple_log_fn(f"{date_str}: sin datos disponibles.")

    finally:
        try:
            driver.quit()
        except Exception:
            pass

    log_fn("")
    log_fn("Resumen exportaciones:")
    for fp, var, action, details in overall_export_results:
        log_fn(f" - {os.path.basename(fp)} | sufijo={var} | {action} | {details}")
    if not overall_export_results:
        log_fn(" - No se encontraron datos en el rango especificado.")

    if remove_html_after:
        removed = 0
        for fname in os.listdir(out_dir):
            if fname.startswith("sbs_fuente_") and fname.endswith(".html"):
                try:
                    os.remove(os.path.join(out_dir, fname))
                    removed += 1
                except Exception:
                    pass
        log_fn(f"[INFO] Eliminados {removed} archivos temporales HTML.")

# ==================== INTERFAZ TKINTER ====================


class TasasApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tasas Pasivas SBS - Rango de fechas")
        self.geometry("840x640")
        self.minsize(840, 480)

        main = ttk.Frame(self, padding=8)
        main.pack(fill="both", expand=True)

        lf_range = ttk.LabelFrame(main, text="Rango de fechas")
        lf_range.pack(fill="x", pady=(0, 6))

        top = ttk.Frame(lf_range)
        top.pack(anchor="w", pady=4, padx=4)

        ttk.Label(top, text="Fecha inicio:").grid(column=0, row=0, sticky="w")
        self.s_day   = ttk.Entry(top, width=3, justify="center", validate="key")
        self.s_month = ttk.Entry(top, width=3, justify="center", validate="key")
        self.s_year  = ttk.Entry(top, width=6, justify="center", validate="key")
        ttk.Label(top, text="/").grid(column=2, row=0)
        ttk.Label(top, text="/").grid(column=4, row=0)
        self.s_day.grid(  column=1, row=0, padx=(2, 2))
        self.s_month.grid(column=3, row=0, padx=(2, 2))
        self.s_year.grid( column=5, row=0, padx=(2, 20))

        ttk.Label(top, text="Fecha fin:").grid(column=6, row=0, sticky="w")
        self.e_day   = ttk.Entry(top, width=3, justify="center", validate="key")
        self.e_month = ttk.Entry(top, width=3, justify="center", validate="key")
        self.e_year  = ttk.Entry(top, width=6, justify="center", validate="key")
        ttk.Label(top, text="/").grid(column=8,  row=0)
        ttk.Label(top, text="/").grid(column=10, row=0)
        self.e_day.grid(  column=7,  row=0, padx=(2, 2))
        self.e_month.grid(column=9,  row=0, padx=(2, 2))
        self.e_year.grid( column=11, row=0, padx=(2, 2))

        ttk.Label(lf_range, text="Formato: DD / MM / AAAA").pack(
            anchor="w", padx=4, pady=(0, 4)
        )

        vcmd_day  = (self.register(self._validate_digits_len), "%P", "2")
        vcmd_year = (self.register(self._validate_digits_len), "%P", "4")
        for w in (self.s_day, self.s_month, self.e_day, self.e_month):
            w.config(validate="key", validatecommand=vcmd_day)
        for w in (self.s_year, self.e_year):
            w.config(validate="key", validatecommand=vcmd_year)

        lf_out     = ttk.LabelFrame(main, text="Salida")
        lf_out.pack(fill="x", pady=(0, 6))
        folder_row = ttk.Frame(lf_out)
        folder_row.pack(fill="x", pady=4)
        ttk.Label(folder_row, text="Carpeta base de salida:").pack(side="left")
        self.out_var   = tk.StringVar(value=os.getcwd())
        self.out_entry = ttk.Entry(folder_row, textvariable=self.out_var, width=60)
        self.out_entry.pack(side="left", padx=6)
        ttk.Button(folder_row, text="Cambiar...", command=self.choose_folder).pack(side="left")

        box = ttk.LabelFrame(main, text="Opciones de importación")
        box.pack(fill="x", pady=6)
        self.var_nat = tk.BooleanVar(value=True)
        self.var_jur = tk.BooleanVar(value=True)
        self.var_gen = tk.BooleanVar(value=True)
        ttk.Checkbutton(box, text="Importar Personas Naturales",           variable=self.var_nat).pack(side="left", padx=6, pady=6)
        ttk.Checkbutton(box, text="Importar Personas Jurídicas",           variable=self.var_jur).pack(side="left", padx=6, pady=6)
        ttk.Checkbutton(box, text="Importar Tasas sin distinguir persona", variable=self.var_gen).pack(side="left", padx=6, pady=6)

        self.adv_btn = ttk.Button(main, text="Opciones avanzadas ▸", command=self.toggle_advanced)
        self.adv_btn.pack(anchor="w")

        self.adv_frame   = ttk.LabelFrame(main, text="Opciones avanzadas")
        adv_opts_row     = ttk.Frame(self.adv_frame)
        adv_opts_row.pack(fill="x", padx=4, pady=4)

        self.var_skip_weekends = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            adv_opts_row, text="Omitir sábados y domingos en el rango",
            variable=self.var_skip_weekends,
        ).pack(side="left", padx=4)

        self.var_remove_html = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            adv_opts_row, text="Eliminar archivos HTML temporales al terminar",
            variable=self.var_remove_html,
        ).pack(side="left", padx=4)

        self.var_show_adv_console = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            adv_opts_row, text="Mostrar consola avanzada",
            variable=self.var_show_adv_console,
            command=self.toggle_advanced_console,
        ).pack(side="left", padx=4)

        self.adv_shown       = False
        self.current_console = "simple"

        self.ctrl_frame = ttk.Frame(main)
        self.ctrl_frame.pack(fill="x", pady=6)
        self.run_btn = ttk.Button(self.ctrl_frame, text="Ejecutar", command=self.on_run)
        self.run_btn.pack(side="left", padx=6)
        self.stop_btn = ttk.Button(
            self.ctrl_frame, text="Detener", command=self.on_stop, state="disabled"
        )
        self.stop_btn.pack(side="left")

        self.console_frame = ttk.LabelFrame(main, text="Consola")
        self.console_frame.pack(fill="both", expand=True, pady=6)

        self.simple_log_widget = scrolledtext.ScrolledText(
            self.console_frame, height=18, state="disabled", wrap="word"
        )
        self.simple_log_widget.pack(fill="both", expand=True)

        self.adv_log_widget = scrolledtext.ScrolledText(
            self.console_frame, height=18, state="disabled", wrap="word"
        )

        today = datetime.now()
        sdate = today - timedelta(days=30)
        self.s_day.insert(0,   sdate.strftime("%d"))
        self.s_month.insert(0, sdate.strftime("%m"))
        self.s_year.insert(0,  sdate.strftime("%Y"))
        self.e_day.insert(0,   today.strftime("%d"))
        self.e_month.insert(0, today.strftime("%m"))
        self.e_year.insert(0,  today.strftime("%Y"))

        self.worker = None

    def _validate_digits_len(self, proposed, maxlen):
        if proposed == "":
            return True
        if not proposed.isdigit():
            return False
        return len(proposed) <= int(maxlen)

    def choose_folder(self):
        d = filedialog.askdirectory(initialdir=self.out_var.get() or os.getcwd())
        if d:
            self.out_var.set(d)

    def toggle_advanced(self):
        if self.adv_shown:
            self.adv_frame.pack_forget()
            self.adv_shown = False
            self.adv_btn.config(text="Opciones avanzadas ▸")
        else:
            self.adv_frame.pack(fill="x", pady=4, before=self.ctrl_frame)
            self.adv_shown = True
            self.adv_btn.config(text="Opciones avanzadas ▾")

    def toggle_advanced_console(self):
        show = self.var_show_adv_console.get()
        if show:
            self.simple_log_widget.pack_forget()
            self.adv_log_widget.pack(fill="both", expand=True)
            self.console_frame.config(text="Consola avanzada")
            self.current_console = "advanced"
        else:
            self.adv_log_widget.pack_forget()
            self.simple_log_widget.pack(fill="both", expand=True)
            self.console_frame.config(text="Consola")
            self.current_console = "simple"

    def log_simple(self, msg: str):
        self.simple_log_widget.config(state="normal")
        self.simple_log_widget.insert("end", msg + "\n")
        self.simple_log_widget.see("end")
        self.simple_log_widget.config(state="disabled")

    def log_advanced(self, msg: str):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.adv_log_widget.config(state="normal")
        self.adv_log_widget.insert("end", f"[{ts}] {msg}\n")
        self.adv_log_widget.see("end")
        self.adv_log_widget.config(state="disabled")

    def on_run(self):
        try:
            s_day  = self.s_day.get().zfill(2)
            s_mon  = self.s_month.get().zfill(2)
            s_year = self.s_year.get().zfill(4)
            e_day  = self.e_day.get().zfill(2)
            e_mon  = self.e_month.get().zfill(2)
            e_year = self.e_year.get().zfill(4)
            start  = f"{s_day}/{s_mon}/{s_year}"
            end    = f"{e_day}/{e_mon}/{e_year}"
            parse_ddmmyyyy(start)
            parse_ddmmyyyy(end)
            if parse_ddmmyyyy(end) < parse_ddmmyyyy(start):
                raise ValueError("end_date < start_date")
        except Exception:
            messagebox.showerror(
                "Fecha inválida",
                "Por favor completa las fechas con valores válidos y verifica "
                "que la fecha fin no sea anterior a la fecha inicio.",
            )
            return

        out_base_dir = self.out_var.get().strip() or os.getcwd()
        self.run_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.log_simple(f"Iniciando proceso desde {start} hasta {end}.")
        self.log_advanced(f"[INFO] Iniciando proceso desde {start} hasta {end}.")

        self.worker = threading.Thread(
            target=self._worker, args=(start, end, out_base_dir), daemon=True
        )
        self.worker.start()
        self.after(300, self._poll_worker)

    def _worker(self, start, end, out_base_dir):
        try:
            run_date_range(
                start, end,
                out_base_dir      = out_base_dir,
                log_fn            = lambda s: self.after(0, lambda: self.log_advanced(s)),
                simple_log_fn     = lambda s: self.after(0, lambda: self.log_simple(s)),
                import_natural    = self.var_nat.get(),
                import_juridica   = self.var_jur.get(),
                import_general    = self.var_gen.get(),
                remove_html_after = self.var_remove_html.get(),
                skip_weekends     = self.var_skip_weekends.get(),
            )
        except Exception as e:
            self.after(0, lambda: self.log_advanced(f"[ERROR] Excepción en worker: {e}"))
            self.after(0, lambda: self.log_advanced(traceback.format_exc()))
            self.after(0, lambda: self.log_simple("Se produjo un error. Revisa la consola avanzada."))

    def _poll_worker(self):
        if self.worker and self.worker.is_alive():
            self.after(500, self._poll_worker)
        else:
            self.run_btn.config(state="normal")
            self.stop_btn.config(state="disabled")
            self.log_simple("Proceso terminado.")
            self.log_advanced("[INFO] Proceso terminado.")

    def on_stop(self):
        self.log_simple("Detención solicitada por el usuario.")
        self.log_advanced(
            "[INFO] Solicitud de detener recibida. "
            "Cierra Chrome manualmente si deseas detener más rápido."
        )
        self.run_btn.config(state="normal")
        self.stop_btn.config(state="disabled")

# ==================== ENTRYPOINT ====================


def main():
    app = TasasApp()
    app.mainloop()


if __name__ == "__main__":
    main()